from flask import Blueprint, render_template, request, redirect, url_for, jsonify, flash, send_file, current_app
from utils.db import get_db_connection, execute_query
from flask_ckeditor import upload_success, upload_fail, CKEditor
import pymysql
import os
import shutil
from datetime import datetime
from werkzeug.utils import secure_filename
import uuid
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io

service_bp = Blueprint('service', __name__, url_prefix='/service')


@service_bp.route('/')
def index():
    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    try:
        cursor.execute("""
            SELECT app_idx, app_servicecode, app_name, app_group, app_domain, app_appcode
            FROM info_service
            ORDER BY app_servicecode
        """)
        services = cursor.fetchall()
    except Exception as e:
        print(f"Error fetching services: {e}")
        services = []
    finally:
        cursor.close()
        conn.close()

    return render_template('service/index.html', services=services)


@service_bp.route('/add', methods=['GET', 'POST'])
def add_service():
    if request.method == 'POST':
        app_servicecode = request.form.get('app_servicecode')
        app_name = request.form.get('app_name')
        app_group = request.form.get('app_group')
        app_domain = request.form.get('app_domain')
        app_appcode = request.form.get('app_appcode')
        app_comment = request.form.get('app_comment', '')

        conn = get_db_connection()
        cursor = conn.cursor()

        try:
            cursor.execute("""
                INSERT INTO info_service (app_servicecode, app_name, app_group, app_domain, app_appcode, app_comment)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (app_servicecode, app_name, app_group, app_domain, app_appcode, app_comment))

            # 새로 생성된 서비스의 app_idx 가져오기
            app_idx = cursor.lastrowid

            # 서비스 파일 디렉토리 생성
            service_dir = os.path.join(current_app.root_path, 'autodata', 'service', f"{app_servicecode}-{app_idx}")
            os.makedirs(service_dir, exist_ok=True)

            conn.commit()
            flash('서비스가 성공적으로 추가되었습니다.', 'success')
            return redirect(url_for('service.index'))
        except Exception as e:
            conn.rollback()
            flash(f'서비스 추가 중 오류가 발생했습니다: {e}', 'danger')
        finally:
            cursor.close()
            conn.close()

    return render_template('service/add.html')


@service_bp.route('/edit/<int:app_idx>', methods=['GET', 'POST'])
def edit_service(app_idx):
    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    if request.method == 'POST':
        app_servicecode = request.form.get('app_servicecode')
        app_name = request.form.get('app_name')
        app_group = request.form.get('app_group')
        app_domain = request.form.get('app_domain')
        app_appcode = request.form.get('app_appcode')
        app_comment = request.form.get('app_comment', '')

        try:
            # 기존 서비스 코드 가져오기
            cursor.execute("SELECT app_servicecode FROM info_service WHERE app_idx = %s", (app_idx,))
            old_service = cursor.fetchone()
            old_servicecode = old_service['app_servicecode'] if old_service else None

            # 서비스 정보 업데이트
            cursor.execute("""
                UPDATE info_service
                SET app_servicecode = %s, app_name = %s, app_group = %s, app_domain = %s, app_appcode = %s, app_comment = %s
                WHERE app_idx = %s
            """, (app_servicecode, app_name, app_group, app_domain, app_appcode, app_comment, app_idx))

            # 서비스 코드가 변경된 경우 디렉토리 이름 변경
            if old_servicecode and old_servicecode != app_servicecode:
                old_dir = os.path.join(current_app.root_path, 'autodata', 'service', f"{old_servicecode}-{app_idx}")
                new_dir = os.path.join(current_app.root_path, 'autodata', 'service', f"{app_servicecode}-{app_idx}")

                if os.path.exists(old_dir):
                    os.makedirs(os.path.dirname(new_dir), exist_ok=True)
                    shutil.move(old_dir, new_dir)

            conn.commit()
            flash('서비스가 성공적으로 수정되었습니다.', 'success')
            return redirect(url_for('service.index'))
        except Exception as e:
            conn.rollback()
            flash(f'서비스 수정 중 오류가 발생했습니다: {e}', 'danger')

    try:
        # 서비스 정보 가져오기
        cursor.execute("""
            SELECT app_idx, app_servicecode, app_name, app_group, app_domain, app_appcode, app_comment
            FROM info_service
            WHERE app_idx = %s
        """, (app_idx,))
        service = cursor.fetchone()

        if not service:
            flash('해당 서비스를 찾을 수 없습니다.', 'danger')
            return redirect(url_for('service.index'))

        # 연계된 자산 목록 가져오기 - 필요한 필드만 조회
        cursor.execute("""
            SELECT a.pnum, a.servername, a.hostname, a.ip, 
            CASE a.domain 
                WHEN 0 THEN '서버' 
                WHEN 1 THEN '스위치' 
                WHEN 2 THEN '스토리지' 
                WHEN 3 THEN '어플라이언스'
                ELSE '미설정' 
            END as domain_state
            FROM total_asset a
            JOIN total_service ts ON a.pnum = ts.service_pnum
            WHERE ts.service_appidx = %s
        """, (app_idx,))
        linked_assets = cursor.fetchall()

        # 서비스 파일 목록 가져오기
        service_files = []
        service_dir = os.path.join(current_app.root_path, 'autodata', 'service',
                                   f"{service['app_servicecode']}-{app_idx}")

        if os.path.exists(service_dir):
            for filename in os.listdir(service_dir):
                file_path = os.path.join(service_dir, filename)
                if os.path.isfile(file_path):
                    file_size = os.path.getsize(file_path)
                    file_size_str = format_file_size(file_size)
                    file_mtime = datetime.fromtimestamp(os.path.getmtime(file_path))

                    service_files.append({
                        'filename': filename,
                        'size': file_size_str,
                        'upload_date': file_mtime.strftime('%Y-%m-%d %H:%M:%S')
                    })

    except Exception as e:
        flash(f'서비스 정보를 가져오는 중 오류가 발생했습니다: {e}', 'danger')
        return redirect(url_for('service.index'))
    finally:
        cursor.close()
        conn.close()

    return render_template('service/edit.html', service=service, linked_assets=linked_assets,
                           service_files=service_files)


@service_bp.route('/delete/<int:app_idx>', methods=['POST'])
def delete_service(app_idx):
    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    try:
        # 서비스 코드 가져오기
        cursor.execute("SELECT app_servicecode FROM info_service WHERE app_idx = %s", (app_idx,))
        service = cursor.fetchone()

        if service:
            # 서비스 디렉토리 삭제
            service_dir = os.path.join(current_app.root_path, 'autodata', 'service',
                                       f"{service['app_servicecode']}-{app_idx}")
            if os.path.exists(service_dir):
                shutil.rmtree(service_dir)

        # 먼저 연계 정보 삭제
        cursor.execute("DELETE FROM total_service WHERE service_appidx = %s", (app_idx,))

        # 서비스 정보 삭제
        cursor.execute("DELETE FROM info_service WHERE app_idx = %s", (app_idx,))
        conn.commit()
        flash('서비스가 성공적으로 삭제되었습니다.', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'서비스 삭제 중 오류가 발생했습니다: {e}', 'danger')
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('service.index'))


@service_bp.route('/linked_services/<int:pnum>')
def get_linked_services(pnum):
    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    try:
        cursor.execute("""
            SELECT s.app_idx, s.app_servicecode, s.app_name, s.app_group, s.app_domain, s.app_appcode
            FROM info_service s
            JOIN total_service ts ON s.app_idx = ts.service_appidx
            WHERE ts.service_pnum = %s
        """, (pnum,))
        services = cursor.fetchall()
        return jsonify(services)
    except Exception as e:
        print(f"Error fetching linked services: {e}")
        return jsonify([])
    finally:
        cursor.close()
        conn.close()


@service_bp.route('/link_asset', methods=['POST'])
def link_asset():
    app_idx = request.form.get('app_idx', type=int)
    pnum = request.form.get('pnum', type=int)

    if not app_idx or not pnum:
        return jsonify({'success': False, 'message': '필수 파라미터가 누락되었습니다.'})

    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    try:
        # 이미 연계되어 있는지 확인
        cursor.execute("""
            SELECT COUNT(*) ct FROM total_service
            WHERE service_appidx = %s AND service_pnum = %s
        """, (app_idx, pnum))
        if cursor.fetchone()['ct'] > 0:
            return jsonify({'success': False, 'message': '이미 연계된 자산입니다.'})
        # 연계 정보 추가
        cursor.execute("""
            INSERT INTO total_service (service_appidx, service_pnum)
            VALUES (%s, %s)
        """, (app_idx, pnum))
        conn.commit()
        return jsonify({'success': True, 'message': '자산이 성공적으로 연계되었습니다.'})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)})
    finally:
        cursor.close()
        conn.close()


@service_bp.route('/unlink_asset', methods=['POST'])
def unlink_asset():
    app_idx = request.form.get('app_idx')
    pnum = request.form.get('pnum')

    if not app_idx or not pnum:
        return jsonify({'success': False, 'message': '필수 파라미터가 누락되었습니다.'})

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            DELETE FROM total_service
            WHERE service_appidx = %s AND service_pnum = %s
        """, (app_idx, pnum))
        conn.commit()
        return jsonify({'success': True, 'message': '자산 연계가 성공적으로 해제되었습니다.'})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)})
    finally:
        cursor.close()
        conn.close()


@service_bp.route('/link_assets', methods=['POST'])
def link_assets():
    app_idx = request.form.get('app_idx')
    asset_pnums = request.form.getlist('asset_pnums[]')

    print(f"자산 연계 요청: app_idx = {app_idx}, asset_pnums = {asset_pnums}")

    if not app_idx or not asset_pnums:
        print("필수 파라미터 누락")
        return jsonify({'success': False, 'message': '필수 파라미터가 누락되었습니다.'})

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        success_count = 0
        already_linked = 0

        for pnum in asset_pnums:
            print(f"자산 {pnum} 연계 처리 중...")
            # 이미 연계되어 있는지 확인
            cursor.execute("""
                SELECT COUNT(*) FROM total_service
                WHERE service_appidx = %s AND service_pnum = %s
            """, (app_idx, pnum))

            if cursor.fetchone()[0] > 0:
                print(f"자산 {pnum}은 이미 연계되어 있음")
                already_linked += 1
                continue

            # 연계 정보 추가
            cursor.execute("""
                INSERT INTO total_service (service_appidx, service_pnum)
                VALUES (%s, %s)
            """, (app_idx, pnum))
            print(f"자산 {pnum} 연계 성공")
            success_count += 1

        conn.commit()

        message = f"{success_count}개 자산이 성공적으로 연계되었습니다."
        if already_linked > 0:
            message += f" ({already_linked}개 자산은 이미 연계되어 있습니다.)"

        print(f"연계 처리 결과: {message}")
        return jsonify({'success': True, 'message': message})
    except Exception as e:
        conn.rollback()
        print(f"자산 연계 중 오류 발생: {e}")
        return jsonify({'success': False, 'message': str(e)})
    finally:
        cursor.close()
        conn.close()


@service_bp.route('/search_assets')
def search_assets():
    term = request.args.get('term', '')
    print(f"자산 검색 요청: 검색어 = '{term}'")

    if not term:
        print("검색어가 비어 있습니다.")
        return jsonify([])

    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    try:
        # 검색 쿼리 수정 - 필요한 필드만 조회하고 도메인 필터링 제거
        query = """
            SELECT a.pnum, a.servername, a.hostname, a.ip,
            CASE a.domain 
                WHEN 0 THEN '서버' 
                WHEN 1 THEN '스위치' 
                WHEN 2 THEN '스토리지' 
                WHEN 3 THEN '어플라이언스'
                ELSE '미설정' 
            END as domain_state
            FROM total_asset a
            WHERE a.servername LIKE %s OR a.hostname LIKE %s OR a.ip LIKE %s
            LIMIT 50
        """
        params = (f'%{term}%', f'%{term}%', f'%{term}%')

        cursor.execute(query, params)
        assets = cursor.fetchall()

        # 결과 확인을 위한 로그
        if assets:
            print(f"첫 번째 결과: {assets[0]}")
        else:
            print("검색 결과가 없습니다.")

        return jsonify(assets)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify([])
    finally:
        cursor.close()
        conn.close()


@service_bp.route('/upload_file/<int:app_idx>', methods=['POST'])
def upload_file(app_idx):
    if 'file' not in request.files:
        flash('파일이 없습니다.', 'danger')
        return redirect(url_for('service.edit_service', app_idx=app_idx))

    file = request.files['file']

    if file.filename == '':
        flash('선택된 파일이 없습니다.', 'danger')
        return redirect(url_for('service.edit_service', app_idx=app_idx))

    try:
        # 서비스 코드 가져오기
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        cursor.execute("SELECT app_servicecode FROM info_service WHERE app_idx = %s", (app_idx,))
        service = cursor.fetchone()
        cursor.close()
        conn.close()

        if not service:
            flash('서비스를 찾을 수 없습니다.', 'danger')
            return redirect(url_for('service.index'))

        # 서비스 디렉토리 경로 확인 및 생성
        base_dir = os.path.join(current_app.root_path, 'autodata', 'service')
        if not os.path.exists(base_dir):
            os.makedirs(base_dir, exist_ok=True)

        service_dir = os.path.join(base_dir, f"{service['app_servicecode']}-{app_idx}")
        if not os.path.exists(service_dir):
            os.makedirs(service_dir, exist_ok=True)

        # 파일 이름 보안 처리
        filename = secure_filename(file.filename)
        file_path = os.path.join(service_dir, filename)

        # 파일 저장
        file.save(file_path)

        flash('파일이 성공적으로 업로드되었습니다.', 'success')

    except Exception as e:
        flash(f'파일 업로드 중 오류가 발생했습니다: {e}', 'danger')

    return redirect(url_for('service.edit_service', app_idx=app_idx))


@service_bp.route('/download_file/<int:app_idx>/<path:filename>')
def download_file(app_idx, filename):
    try:
        # 서비스 코드 가져오기
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        cursor.execute("SELECT app_servicecode FROM info_service WHERE app_idx = %s", (app_idx,))
        service = cursor.fetchone()
        cursor.close()
        conn.close()

        if not service:
            flash('서비스를 찾을 수 없습니다.', 'danger')
            return redirect(url_for('service.index'))

        # 파일 경로
        service_dir = os.path.join(current_app.root_path, 'autodata', 'service',
                                   f"{service['app_servicecode']}-{app_idx}")
        file_path = os.path.join(service_dir, filename)

        if not os.path.exists(file_path):
            flash('파일을 찾을 수 없습니다.', 'danger')
            return redirect(url_for('service.edit_service', app_idx=app_idx))

        return send_file(file_path, as_attachment=True)

    except Exception as e:
        flash(f'파일 다운로드 중 오류가 발생했습니다: {e}', 'danger')
        return redirect(url_for('service.edit_service', app_idx=app_idx))


@service_bp.route('/delete_file/<int:app_idx>', methods=['POST'])
def delete_file(app_idx):
    filename = request.form.get('filename')

    if not filename:
        return jsonify({'success': False, 'message': '파일 이름이 제공되지 않았습니다.'})

    try:
        # 서비스 코드 가져오기
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        cursor.execute("SELECT app_servicecode FROM info_service WHERE app_idx = %s", (app_idx,))
        service = cursor.fetchone()
        cursor.close()
        conn.close()

        if not service:
            return jsonify({'success': False, 'message': '서비스를 찾을 수 없습니다.'})

        # 파일 경로
        service_dir = os.path.join(current_app.root_path, 'autodata', 'service',
                                   f"{service['app_servicecode']}-{app_idx}")
        file_path = os.path.join(service_dir, filename)

        if not os.path.exists(file_path):
            return jsonify({'success': False, 'message': '파일을 찾을 수 없습니다.'})

        # 파일 삭제
        os.remove(file_path)

        return jsonify({'success': True, 'message': '파일이 성공적으로 삭제되었습니다.'})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# 서비스 일괄 추가 템플릿 다운로드
@service_bp.route('/download_service_template')
def download_service_template():
    # 엑셀 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "서비스 일괄 추가"

    # 헤더 설정
    headers = ["서비스 코드*", "서비스명*", "서비스 구분", "서비스 도메인", "APP 코드", "서비스 코멘트"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # 열 너비 설정
    for col_num, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    # 예시 데이터 추가
    example_data = [
        ["SVC001", "웹 서비스", "웹", "example.com", "APP001", "웹 서비스 설명"],
        ["SVC002", "모바일 서비스", "모바일", "m.example.com", "APP002", "모바일 서비스 설명"]
    ]

    for row_num, row_data in enumerate(example_data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num).value = cell_value

    # 안내 시트 추가
    ws_guide = wb.create_sheet(title="작성 가이드")
    guide_text = [
        ["서비스 일괄 추가 가이드"],
        [""],
        ["1. '서비스 일괄 추가' 시트에 서비스 정보를 입력하세요."],
        ["2. '*'가 표시된 항목은 필수 입력 항목입니다."],
        ["3. 서비스 코드는 중복되지 않아야 합니다."],
        ["4. 서비스 코멘트는 간단한 텍스트만 입력 가능합니다."],
        ["5. 작성이 완료되면 파일을 저장하고 업로드하세요."]
    ]

    for row_num, row_data in enumerate(guide_text, 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws_guide.cell(row=row_num, column=col_num)
            cell.value = cell_value
            if row_num == 1:
                cell.font = Font(bold=True, size=14)

    # 열 너비 설정
    ws_guide.column_dimensions['A'].width = 50

    # 메모리에 엑셀 파일 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="서비스_일괄추가_템플릿.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# 서비스-자산 연계 템플릿 다운로드
@service_bp.route('/download_link_template')
def download_link_template():
    # 엑셀 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "서비스-자산 연계"

    # 헤더 설정
    headers = ["서비스 코드*", "자산 번호*"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # 열 너비 설정
    for col_num, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    # 예시 데이터 추가
    example_data = [
        ["SVC001", "1001"],
        ["SVC001", "1002"],
        ["SVC002", "1003"]
    ]

    for row_num, row_data in enumerate(example_data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num).value = cell_value

    # 서비스 목록 시트 추가
    ws_services = wb.create_sheet(title="서비스 목록")

    # 서비스 목록 헤더
    service_headers = ["서비스 코드", "서비스명"]
    for col_num, header in enumerate(service_headers, 1):
        cell = ws_services.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # 서비스 목록 데이터 가져오기
    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    try:
        cursor.execute("SELECT app_servicecode, app_name FROM info_service ORDER BY app_servicecode")
        services = cursor.fetchall()

        for row_num, service in enumerate(services, 2):
            ws_services.cell(row=row_num, column=1).value = service['app_servicecode']
            ws_services.cell(row=row_num, column=2).value = service['app_name']
    except Exception as e:
        print(f"Error fetching services for template: {e}")
    finally:
        cursor.close()
        conn.close()

    # 열 너비 설정
    for col_num, header in enumerate(service_headers, 1):
        ws_services.column_dimensions[get_column_letter(col_num)].width = 20

    # 안내 시트 추가
    ws_guide = wb.create_sheet(title="작성 가이드")
    guide_text = [
        ["서비스-자산 연계 가이드"],
        [""],
        ["1. '서비스-자산 연계' 시트에 연계 정보를 입력하세요."],
        ["2. '*'가 표시된 항목은 필수 입력 항목입니다."],
        ["3. 서비스 코드는 '서비스 목록' 시트에서 확인할 수 있습니다."],
        ["4. 자산 번호는 자산 관리 시스템에 등록된 자산의 고유 번호(pnum)입니다."],
        ["5. 하나의 서비스에 여러 자산을 연계할 수 있습니다."],
        ["6. 이미 연계된 서비스-자산 조합은 무시됩니다."],
        ["7. 작성이 완료되면 파일을 저장하고 업로드하세요."]
    ]

    for row_num, row_data in enumerate(guide_text, 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws_guide.cell(row=row_num, column=col_num)
            cell.value = cell_value
            if row_num == 1:
                cell.font = Font(bold=True, size=14)

    # 열 너비 설정
    ws_guide.column_dimensions['A'].width = 60

    # 메모리에 엑셀 파일 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="서비스_자산연계_템플릿.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# 서비스 일괄 추가 처리
@service_bp.route('/upload_services', methods=['POST'])
def upload_services():
    if 'file' not in request.files:
        flash('파일이 없습니다.', 'danger')
        return redirect(url_for('service.index'))

    file = request.files['file']

    if file.filename == '':
        flash('선택된 파일이 없습니다.', 'danger')
        return redirect(url_for('service.index'))

    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('엑셀 파일만 업로드 가능합니다.', 'danger')
        return redirect(url_for('service.index'))

    try:
        # 엑셀 파일 읽기
        df = pd.read_excel(file, sheet_name="서비스 일괄 추가")

        # 필수 컬럼 확인
        required_columns = ["서비스 코드*", "서비스명*"]
        for column in required_columns:
            if column not in df.columns:
                flash(f'필수 컬럼 "{column}"이 없습니다.', 'danger')
                return redirect(url_for('service.index'))

        # 컬럼명 매핑
        column_mapping = {
            "서비스 코드*": "app_servicecode",
            "서비스명*": "app_name",
            "서비스 구분": "app_group",
            "서비스 도메인": "app_domain",
            "APP 코드": "app_appcode",
            "서비스 코멘트": "app_comment"
        }

        # 컬럼명 변경
        df = df.rename(columns=column_mapping)

        # NaN 값 처리
        for col in ["app_group", "app_domain", "app_appcode", "app_comment"]:
            if col in df.columns:
                df[col] = df[col].fillna("")

        # 필수 값 확인
        if df["app_servicecode"].isnull().any() or df["app_name"].isnull().any():
            flash('서비스 코드와 서비스명은 필수 입력 항목입니다.', 'danger')
            return redirect(url_for('service.index'))

        # 데이터베이스 연결
        conn = get_db_connection()
        cursor = conn.cursor()

        success_count = 0
        error_count = 0
        duplicate_count = 0

        try:
            # 기존 서비스 코드 가져오기
            cursor.execute("SELECT app_servicecode FROM info_service")
            existing_codes = [row[0] for row in cursor.fetchall()]

            for _, row in df.iterrows():
                try:
                    # 중복 서비스 코드 확인
                    if row["app_servicecode"] in existing_codes:
                        duplicate_count += 1
                        continue

                    # 서비스 추가
                    sql = """
                        INSERT INTO info_service (app_servicecode, app_name, app_group, app_domain, app_appcode, app_comment)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """

                    values = (
                        row["app_servicecode"],
                        row["app_name"],
                        row.get("app_group", ""),
                        row.get("app_domain", ""),
                        row.get("app_appcode", ""),
                        row.get("app_comment", "")
                    )

                    cursor.execute(sql, values)

                    # 새로 생성된 서비스의 app_idx 가져오기
                    app_idx = cursor.lastrowid

                    # 서비스 파일 디렉토리 생성
                    service_dir = os.path.join(current_app.root_path, 'autodata', 'service',
                                               f"{row['app_servicecode']}-{app_idx}")
                    os.makedirs(service_dir, exist_ok=True)

                    success_count += 1
                    existing_codes.append(row["app_servicecode"])  # 추가된 코드를 기존 코드 목록에 추가

                except Exception as e:
                    print(f"Error adding service: {e}")
                    error_count += 1

            conn.commit()

            message = f"{success_count}개 서비스가 성공적으로 추가되었습니다."
            if duplicate_count > 0:
                message += f" {duplicate_count}개 서비스는 중복된 코드로 무시되었습니다."
            if error_count > 0:
                message += f" {error_count}개 서비스는 오류로 추가되지 않았습니다."

            flash(message, 'success' if success_count > 0 else 'warning')

        except Exception as e:
            conn.rollback()
            flash(f'서비스 일괄 추가 중 오류가 발생했습니다: {e}', 'danger')
        finally:
            cursor.close()
            conn.close()

    except Exception as e:
        flash(f'엑셀 파일 처리 중 오류가 발생했습니다: {e}', 'danger')

    return redirect(url_for('service.index'))


# 서비스-자산 연계 일괄 처리
@service_bp.route('/upload_links', methods=['POST'])
def upload_links():
    if 'file' not in request.files:
        flash('파일이 없습니다.', 'danger')
        return redirect(url_for('service.index'))

    file = request.files['file']

    if file.filename == '':
        flash('선택된 파일이 없습니다.', 'danger')
        return redirect(url_for('service.index'))

    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('엑셀 파일만 업로드 가능합니다.', 'danger')
        return redirect(url_for('service.index'))

    try:
        # 엑셀 파일 읽기
        df = pd.read_excel(file, sheet_name="서비스-자산 연계")

        # 필수 컬럼 확인
        required_columns = ["서비스 코드*", "자산 번호*"]
        for column in required_columns:
            if column not in df.columns:
                flash(f'필수 컬럼 "{column}"이 없습니다.', 'danger')
                return redirect(url_for('service.index'))

        # 컬럼명 매핑
        column_mapping = {
            "서비스 코드*": "app_servicecode",
            "자산 번호*": "pnum"
        }

        # 컬럼명 변경
        df = df.rename(columns=column_mapping)

        # 필수 값 확인
        if df["app_servicecode"].isnull().any() or df["pnum"].isnull().any():
            flash('서비스 코드와 자산 번호는 필수 입력 항목입니다.', 'danger')
            return redirect(url_for('service.index'))

        # 데이터베이스 연결
        conn = get_db_connection()
        cursor = conn.cursor()

        success_count = 0
        error_count = 0
        already_linked_count = 0
        invalid_count = 0

        try:
            for _, row in df.iterrows():
                try:
                    # 서비스 코드로 app_idx 가져오기
                    cursor.execute("SELECT app_idx FROM info_service WHERE app_servicecode = %s",
                                   (row["app_servicecode"],))
                    service_result = cursor.fetchone()

                    if not service_result:
                        invalid_count += 1
                        continue

                    app_idx = service_result[0]
                    pnum = int(row["pnum"])

                    # 자산 번호 유효성 확인
                    cursor.execute("SELECT COUNT(*) FROM total_asset WHERE pnum = %s", (pnum,))
                    if cursor.fetchone()[0] == 0:
                        invalid_count += 1
                        continue

                    # 이미 연계되어 있는지 확인
                    cursor.execute("""
                        SELECT COUNT(*) FROM total_service
                        WHERE service_appidx = %s AND service_pnum = %s
                    """, (app_idx, pnum))

                    if cursor.fetchone()[0] > 0:
                        already_linked_count += 1
                        continue

                    # 연계 정보 추가
                    cursor.execute("""
                        INSERT INTO total_service (service_appidx, service_pnum)
                        VALUES (%s, %s)
                    """, (app_idx, pnum))

                    success_count += 1

                except Exception as e:
                    print(f"Error linking service-asset: {e}")
                    error_count += 1

            conn.commit()

            message = f"{success_count}개 서비스-자산 연계가 성공적으로 추가되었습니다."
            if already_linked_count > 0:
                message += f" {already_linked_count}개는 이미 연계되어 있어 무시되었습니다."
            if invalid_count > 0:
                message += f" {invalid_count}개는 유효하지 않은 서비스 코드 또는 자산 번호로 무시되었습니다."
            if error_count > 0:
                message += f" {error_count}개는 오류로 추가되지 않았습니다."

            flash(message, 'success' if success_count > 0 else 'warning')

        except Exception as e:
            conn.rollback()
            flash(f'서비스-자산 연계 일괄 처리 중 오류가 발생했습니다: {e}', 'danger')
        finally:
            cursor.close()
            conn.close()

    except Exception as e:
        flash(f'엑셀 파일 처리 중 오류가 발생했습니다: {e}', 'danger')

    return redirect(url_for('service.index'))


# 유틸리티 함수
def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions


def format_file_size(size_bytes):
    """파일 크기를 읽기 쉬운 형식으로 변환"""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    elif size_bytes < 1024 * 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.1f} MB"
    else:
        return f"{size_bytes / (1024 * 1024 * 1024):.1f} GB"
