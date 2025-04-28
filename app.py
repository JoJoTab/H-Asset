from flask import Flask, render_template, redirect, url_for, request, send_file, jsonify, session, flash
import pandas as pd
import numpy as np
import pymysql.cursors
import os
from datetime import datetime, timedelta
from flask_wtf import FlaskForm
from flask_wtf.file import FileField, FileRequired
import openpyxl
import plotly.express as px
import asyncio
import os
from datetime import datetime, timedelta
from flask_wtf import FlaskForm
from flask_wtf.file import FileField, FileRequired
import aiofiles
from flask_cors import CORS
import redis
from config import Config
from blueprints.asset import asset_bp
from blueprints.storage import storage_bp
from blueprints.rack import rack_bp
from blueprints.file import file_bp
from blueprints.trend import trend_bp
from utils.db import init_db_pool, close_db_pool
from utils.auto_register import setup_auto_register
from utils.auto_storage import setup_auto_storage

app = Flask(__name__)
app.config.from_object(Config)

# 블루프린트 등록
app.register_blueprint(asset_bp)
app.register_blueprint(storage_bp)
app.register_blueprint(rack_bp)
app.register_blueprint(file_bp)
app.register_blueprint(trend_bp)

# 자동 등록 기능 설정
setup_auto_register()
setup_auto_storage()

# 비동기 함수 호출 제거
@app.before_first_request
def setup():
    # 비동기 함수를 동기 함수로 변경
    init_db_pool()
    print("애플리케이션 초기화 완료")

@app.teardown_appcontext
def teardown(exception):
    # 비동기 함수를 동기 함수로 변경
    close_db_pool()
    print("애플리케이션 종료 완료")

@app.route('/')
@app.route('/index')
def main_index():
    # 비동기 함수를 직접 호출하지 않고 동기 버전의 함수 호출
    return redirect(url_for('asset.index'))

def get_db_connection():
    return pymysql.connect(
        host="localhost",
        user="root",
        passwd="a980911",
        db="hli_asset",
        charset="utf8",
        cursorclass=pymysql.cursors.DictCursor
    )


def get_data():
    # 데이터베이스 연결
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            # 전체 데이터 조회
            query = "SELECT * FROM total_asset"
            cursor.execute(query)
            data = cursor.fetchall()
            columns = [column[0] for column in cursor.description]
            df = pd.DataFrame(data, columns=columns)

            # 날짜 형식 변환 (오류를 피하기 위해 errors='coerce' 사용)
            df['datein'] = pd.to_datetime(df['datein'], errors='coerce')
    finally:
        connection.close()

    # 개수 계산
    total_assets = df[df['isoper'].isin([0, 1, 2])]
    total_servers = df[total_assets['domain'] == 0]
    print(len(total_servers), len(df), len(total_servers['isvm']))
    physical_servers = df[total_servers['isvm'] == 0]
    virtual_servers = df[total_servers['isvm'] == 1]
    current_year_assets = df[pd.to_datetime(total_assets['datein']).dt.year == pd.to_datetime('now').year]
    current_month_assets = df[pd.to_datetime(total_assets['datein']).dt.month == pd.to_datetime('now').month]
    oper_assets = df[total_servers['oper'] == 0]
    qa_assets = df[total_servers['oper'] == 1]
    dev_assets = df[total_servers['oper'] == 2]
    dr_assets = df[total_servers['oper'] == 4]
    current_year = pd.to_datetime('now').year
    current_month = pd.to_datetime('now').month

    return {
        "total_assets": total_assets.shape[0],
        "total_servers": total_servers.shape[0],
        "physical_servers": physical_servers.shape[0],
        "virtual_servers": virtual_servers.shape[0],
        "current_year_assets": current_year_assets.shape[0],
        "current_month_assets": current_month_assets.shape[0],
        "oper_assets": oper_assets.shape[0],
        "qa_assets": qa_assets.shape[0],
        "dev_assets": dev_assets.shape[0],
        "dr_assets": dr_assets.shape[0],
        "current_year": current_year,
        "current_month": current_month
    }

def insert_data(date, storage_type, pid, av_cap, tp_cap, tl_cap):
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            sql = "INSERT INTO total_storage (DATEIN, STORAGE, PID, AV_CAP, TP_CAP, TL_CAP) VALUES (%s, %s, %s, %s, %s, %s)"
            cursor.execute(sql, (date, storage_type, pid, av_cap, tp_cap, tl_cap))
        connection.commit()
    finally:
        connection.close()

@app.route('/storage', methods=['GET', 'POST'])
def storage():
    connection = get_db_connection()
    data1 = {}
    data2 = []
    end_date = None
    start_date = None
    error_message = None
    try:
        with connection.cursor() as cursor:
            # 전체 데이터 조회
            sql = "SELECT DATEIN, STORAGE, PID, AV_CAP, TP_CAP, TL_CAP FROM total_storage"

            cursor.execute(sql)
            result = cursor.fetchall()

            # 장 최근 날짜 가져오기
            sql_recent_date = """SELECT MAX(DATEIN) AS end_date FROM total_storage"""
            sql_yesterday_date = """SELECT DATE_SUB(MAX(DATEIN), INTERVAL 1 DAY) AS start_date FROM total_storage"""
            cursor.execute(sql_recent_date)
            end_date = cursor.fetchone()['end_date']
            cursor.execute(sql_yesterday_date)
            start_date = cursor.fetchone()['start_date']

            if request.method == 'POST':
                start_date = request.form.get('start_date')
                end_date = request.form.get('end_date')

            print(start_date, end_date)
            # 데이터 조회
            sql_data = """
                        SELECT STORAGE, PID, DATEIN, AV_CAP, TP_CAP, TL_CAP 
                        FROM total_storage 
                        WHERE DATEIN BETWEEN %s AND %s
                        """

            cursor.execute(sql_data, (start_date, end_date))
            date_range_data = cursor.fetchall()

            # 데이터 조회 후 추가
            # SQL 쿼리로 데이터 조회
            sql_data = """
                        SELECT STORAGE, PID, DATEIN, AV_CAP, TP_CAP, TL_CAP 
                        FROM total_storage 
                        WHERE DATEIN BETWEEN %s AND %s
                        """

            cursor.execute(sql_data, (start_date, end_date))
            date_range_data = cursor.fetchall()

            # 스토리지별로 날짜별 데이터를 저장할 딕셔너리 초기화
            data_for_plot = {}

            for row in date_range_data:
                date = row['DATEIN']
                storage = row['STORAGE']

                # 할당률 및 사용률 계산
                tl_rate = row['TL_CAP'] * 100 / row['TP_CAP'] if row['TP_CAP'] > 0 else 0
                use_rate = (row['TP_CAP'] - row['AV_CAP']) * 100 / row['TL_CAP'] if row['TL_CAP'] > 0 else 0

                # 날짜별로 스토리지 데이터를 누적
                if storage not in data_for_plot:
                    data_for_plot[storage] = {}

                if date not in data_for_plot[storage]:
                    data_for_plot[storage][date] = {
                        'tl_rates': [],
                        'use_rates': []
                    }

                data_for_plot[storage][date]['tl_rates'].append(tl_rate)
                data_for_plot[storage][date]['use_rates'].append(use_rate)

            # 최종 데이터 구조화
            final_data_for_plot = {'dates': [], 'storages': [], 'tl_rates': [], 'use_rates': []}
            for storage, date_data in data_for_plot.items():
                for date, rates in date_data.items():
                    avg_tl_rate = sum(rates['tl_rates']) / len(rates['tl_rates']) if rates['tl_rates'] else 0
                    avg_use_rate = sum(rates['use_rates']) / len(rates['use_rates']) if rates['use_rates'] else 0

                    final_data_for_plot['dates'].append(date)
                    final_data_for_plot['storages'].append(storage)
                    final_data_for_plot['tl_rates'].append(avg_tl_rate)
                    final_data_for_plot['use_rates'].append(avg_use_rate)

            # DataFrame 생성
            df = pd.DataFrame(final_data_for_plot)

            # 할당률 꺾은선 그래프 생성
            fig_tl = px.line(df, x='dates', y='tl_rates', color='storages',
                             labels={'tl_rates': '할당률 (%)', 'dates': '날짜'})

            # 사용률 꺾은선 그래프 생성
            fig_use = px.line(df, x='dates', y='use_rates', color='storages',
                              labels={'use_rates': '사용률 (%)', 'dates': '날짜'})

            # 그래프 HTML 코드로 변환
            graph_html_tl = fig_tl.to_html(full_html=False)
            graph_html_use = fig_use.to_html(full_html=False)

            if not date_range_data:
                # 시작 날짜와 종료 날짜의 데이터 유무 검사
                sql_check_data = """SELECT DATEIN FROM total_storage WHERE DATEIN IN (%s, %s)"""
                cursor.execute(sql_check_data, (start_date, end_date))
                existing_dates = {row['DATEIN'] for row in cursor.fetchall()}

                # if start_date not in existing_dates:
                #     error_message = '시작 날짜에 데이터가 없습니다.'
                if end_date not in existing_dates:
                    error_message = '종료 날짜에 데이터가 없습니다.'
            else:
                # 데이터 비교 로직
                filtered_data = [row for row in date_range_data if row['DATEIN'].strftime('%Y-%m-%d') == end_date]
                for recent in filtered_data:
                    storage = recent['STORAGE']
                    pid = recent['PID']
                    if storage not in data1:
                        data1[storage] = {}

                    data1[storage][pid] = {
                        'AV_CAP': recent['AV_CAP'] / 1024 / 1024,  # TB로 변환
                        'TP_CAP': recent['TP_CAP'] / 1024 / 1024,
                        'TL_CAP': recent['TL_CAP'] / 1024 / 1024,
                        'AV_CAP_diff': None,  # 초기화
                        'TP_CAP_diff': None,
                        'TL_CAP_diff': None,
                        'TL_RATE': recent['TL_CAP'] * 100 / recent['TP_CAP'] if recent['TP_CAP'] > 0 else 0,
                        'USE_RATE': (recent['TP_CAP'] - recent['AV_CAP']) * 100 / recent['TL_CAP'] if recent['TL_CAP'] > 0 else 0
                    }
                filtered_data = [row for row in date_range_data if row['DATEIN'].strftime('%Y-%m-%d') == start_date]
                for yesterday in filtered_data:
                    storage = yesterday['STORAGE']
                    pid = yesterday['PID']
                    if storage in data1 and pid in data1[storage]:
                        # 차이 계산
                        diff = round(data1[storage][pid]['AV_CAP'] - (yesterday['AV_CAP'] / 1024 / 1024), 2)
                        data1[storage][pid]['AV_CAP_diff'] = f"({diff:+})"  # 차이를 포맷팅
                        diff = round(data1[storage][pid]['TP_CAP'] - (yesterday['TP_CAP'] / 1024 / 1024), 2)
                        data1[storage][pid]['TP_CAP_diff'] = f"({diff:+})"  # 차이를 포맷팅
                        diff = round(data1[storage][pid]['TL_CAP'] - (yesterday['TL_CAP'] / 1024 / 1024), 2)
                        data1[storage][pid]['TL_CAP_diff'] = f"({diff:+})"  # 차이를 포맷팅

            # MB를 TB로 변환하여 결과에 추가
            for row in result:
                row['AV_CAP'] = round(row['AV_CAP'] / 1024 / 1024, 2)  # TB 단위로 변환
                row['TP_CAP'] = round(row['TP_CAP'] / 1024 / 1024, 2)
                row['TL_CAP'] = round(row['TL_CAP'] / 1024 / 1024, 2)
                data2.append(row)

    finally:
        connection.close()

    return render_template('storage.html', data1=data1, data2=data2, latest_date=end_date,
                           start_date=start_date, end_date=end_date, error_message=error_message,
                           graph_html_tl=graph_html_tl, graph_html_use=graph_html_use)


@app.route('/storage_upload', methods=['POST'])
def storage_upload():
    if 'file' not in request.files:
        return 'No file part', 400

    file = request.files['file']
    if file.filename == '':
        return 'No selected file', 400

    content = file.read().decode('utf-8').splitlines()
    date = content[0].strip()  # 첫 번째 줄에서 날짜 추출
    storage_type = None

    for line in content[1:]:
        line = line.strip()
        if line.startswith("VSP") or line.startswith("F800") or line.startswith("DR_"):  # 스토리지 종류 확인
            storage_type = line  # 스토리지 종류 저장
        elif line.startswith("PID"):
            continue  # 헤더는 무시
        elif line and storage_type:  # 데이터 줄 처리
            data = line.split()
            if len(data) > 10:  # 유효한 데이터인지 확인
                pid = data[0]
                av_cap = data[3]
                tp_cap = data[4]
                tl_cap = data[10]
                insert_data(date, storage_type, pid, av_cap, tp_cap, tl_cap)

    print(date, storage_type, pid, av_cap, tp_cap, tl_cap)

    # 성공적으로 처리된 후 적절한 응답을 반환
    return redirect(url_for('storage'))

@app.route('/trend_os')
def trend_os():
    sql = """
            SELECT io_os.state AS os
            FROM hli_asset.total_asset ta 
            JOIN hli_asset.info_os io_os ON ta.os = io_os.os
            """
    db = get_db_connection()
    try:
        with db.cursor() as cursor:
            cursor.execute(sql)
            data = cursor.fetchall()
            columns = [column[0] for column in cursor.description]
            df = pd.DataFrame(data, columns=columns)
            cursor.close()
    finally:
        db.close()

    # OS 종류별 자산 수 집계
    os_counts = df['os'].value_counts().reset_index()
    os_counts.columns = ['OS', 'Count']

    # 도넛 그래프 생성
    fig = px.pie(os_counts, values='Count', names='OS', title='OS별 서버수량',
                 hole=0.3)

    graph = fig.to_html(full_html=False)

    return render_template('trend_os.html', graph=graph)

@app.route('/trend_os_date')
def trend_os_date():
    sql_filtered = """
                SELECT ta.datein,
                       io_os.state AS os
                FROM hli_asset.total_asset ta 
                JOIN hli_asset.info_os io_os ON ta.os = io_os.os
                """
    db = get_db_connection()
    try:
        with db.cursor() as cursor:
            cursor.execute(sql_filtered)
            data = cursor.fetchall()
            columns = [column[0] for column in cursor.description]
            df = pd.DataFrame(data, columns=columns)
    finally:
        db.close()

    # 날짜 형식 변환
    df['datein'] = pd.to_datetime(df['datein'])

    # 월별 자산 변화량
    df_grouped = df.groupby([pd.Grouper(key='datein', freq='ME'), 'os']).size().reset_index(name='count')
    fig = px.bar(df_grouped, x='datein', y='count', color='os', title='월간 자산 분포',
                 labels={'count': '자산 수량', 'datein': '월'})

    # 월별 데이터 프레임 생성
    all_months = pd.date_range(start=df_grouped['datein'].min(), end=pd.Timestamp.today(), freq='ME')
    all_os = df_grouped['os'].unique()

    # 모든 조합 생성
    index = pd.MultiIndex.from_product([all_months, all_os], names=['datein', 'os'])
    df_full = pd.DataFrame(index=index).reset_index()

    # 기존 데이터와 결합
    df_full = df_full.merge(df_grouped, on=['datein', 'os'], how='left').fillna(0)

    # 누적 합계 계산
    df_full['cumulative_count'] = df_full.groupby('os')['count'].cumsum()

    # 그래프 생성
    fig3 = px.bar(df_full, x='datein', y='cumulative_count', color='os',
                 title='월간 누적 자산 분포',
                 labels={'cumulative_count': '자산 수량', 'datein': '월'})

    # 연간 자산 변화량
    df_grouped_yearly = df.groupby([pd.Grouper(key='datein', freq='YE'), 'os']).size().reset_index(name='count')
    fig2 = px.bar(df_grouped_yearly, x='datein', y='count', color='os', title='연간 자산 분포',
                 labels={'count': '자산 수량', 'datein': '연도'})

    # 모든 연도 데이터 프레임 생성
    all_years = pd.date_range(start=df_grouped_yearly['datein'].min(), end=pd.Timestamp.today(), freq='YE')

    # 모든 조합 생성
    index_year = pd.MultiIndex.from_product([all_years, all_os], names=['datein', 'os'])
    df_full_year = pd.DataFrame(index=index_year).reset_index()

    # 기존 데이터와 결합
    df_full_year = df_full_year.merge(df_grouped_yearly, on=['datein', 'os'], how='left').fillna(0)

    # 누적 합계 계산
    df_full_year['cumulative_count'] = df_full_year.groupby('os')['count'].cumsum()

    # 그래프 생성
    fig4 = px.bar(df_full_year, x='datein', y='cumulative_count', color='os',
                  title='연간 누적 자산 분포',
                  labels={'cumulative_count': '자산 수량', 'datein': '연도'})

    # HTML로 그래프 렌더링
    graph1 = fig.to_html(full_html=False)
    graph2 = fig2.to_html(full_html=False)
    graph3 = fig3.to_html(full_html=False)
    graph4 = fig4.to_html(full_html=False)

    return render_template('trend_os_date.html', graph1=graph1, graph2=graph2, graph3=graph3, graph4=graph4)


@app.route('/racklayout_edit', methods=['GET', 'POST'])
def racklayout_edit():
    if request.method == 'POST':
        loc = request.form.get('update')  # 수정 버튼을 통해 전달된 loc 값
        if loc:
            rackname = request.form.get(f'rackname_{loc}', '')
            rackenable = int(request.form.get(f'rackenable_{loc}', 1))

            # loc에 해당하는 rack_info 업데이트 또는 삽입
            sql_check = "SELECT COUNT(*) as count FROM rack_info WHERE loc = %s"
            try:
                db = get_db_connection()
                with db.cursor() as cursor:
                    cursor.execute(sql_check, (loc,))
                    result = cursor.fetchone()
                    exists = result['count'] if result else 0  # 결과가 없으면 0으로 설정

                    if exists > 0:
                        # 존재하면 업데이트
                        sql_update = """
                        UPDATE rack_info 
                        SET rackname = %s, rackenable = %s 
                        WHERE loc = %s
                        """
                        cursor.execute(sql_update, (rackname, rackenable, loc))
                    else:
                        # 존재하지 않으면 삽입
                        sql_insert = """
                        INSERT INTO rack_info (loc, rackname, rackenable) 
                        VALUES (%s, %s, %s)
                        """
                        cursor.execute(sql_insert, (loc, rackname, rackenable))

                    db.commit()  # 모든 변경사항을 커밋
            finally:
                db.close()

    # loc1 정보 조회
    sql_loc = "SELECT DISTINCT loc1 FROM total_asset"
    sql_select = "SELECT loc, rackname, rackenable FROM rack_info WHERE loc IN %s"
    try:
        db = get_db_connection()
        with db.cursor() as cursor:
            cursor.execute(sql_loc)
            loc_data = cursor.fetchall()

            # loc_data에서 loc1 값 추출 및 정렬
            loc_list = sorted([row['loc1'] for row in loc_data if row['loc1'] is not None])  # 각 딕셔너리에서 'loc1' 값을 가져와 정렬

            cursor.execute(sql_select, (tuple(loc_list),))
            current_data = cursor.fetchall()
    finally:
        db.close()

    # 데이터 사전 생성
    current_dict = {row['loc']: (row['rackname'], row['rackenable']) for row in current_data}  # 딕셔너리에서 키 사용

    return render_template('racklayout_edit.html', loc_list=loc_list, current_dict=current_dict)


@app.route('/racklayout', methods=['GET'])
def racklayout():
    sql = "SELECT loc1 FROM total_asset WHERE isvm = 0"
    db = get_db_connection()
    try:
        with db.cursor() as cursor:
            cursor.execute(sql)
            data = cursor.fetchall()
    finally:
        db.close()

    df = pd.DataFrame(data, columns=['loc1'])
    df = df.dropna(subset=['loc1'])

    df['loc1'] = df['loc1'].apply(lambda x: x[:-4] + x[-3:-2] + '-' + x[-2:])

    floors = set()
    floor_columns = {}

    for loc in df['loc1']:
        parts = loc.split('-')
        if len(parts) < 3:
            continue  # 형식이 올바르지 않으면 건너뜀

        floor = parts[0]
        column = parts[1]  # R01, L01 등

        floors.add(floor)
        if floor not in floor_columns:
            floor_columns[floor] = set()
        floor_columns[floor].add(column)

    # 랙 정보 가져오기
    db = get_db_connection()
    sql_rack_info = "SELECT loc, rackname, rackenable FROM rack_info"
    try:
        with db.cursor() as cursor:
            cursor.execute(sql_rack_info)
            rack_info_data = cursor.fetchall()
            columns = [column[0] for column in cursor.description]
    finally:
        db.close()
    df_rack = pd.DataFrame(rack_info_data, columns=columns)
    df_rack['loc'] = df_rack['loc'].apply(lambda x: x[:-4] + x[-3:-2] + '-' + x[-2:])
    rack_info_dict = {row['loc']: (row['rackname'], row['rackenable']) for index, row in df_rack.iterrows()}

    # 클릭 가능한 슬롯 생성
    equipment_data = {}
    for floor in floors:
        for column in floor_columns[floor]:
            for loc in range(0, 16):  # 00~15 (0부터 시작)
                loc_string = '{0:02d}'.format(loc)
                loc_key = "{}-{}-{}".format(floor, column, loc_string)
                equipment_data[loc_key] = loc_key in df['loc1'].values
        if floor in floor_columns:
            # R과 L로 분리
            r_items = [item for item in floor_columns[floor] if item.endswith('R')]
            l_items = [item for item in floor_columns[floor] if item.endswith('L')]

            # R과 L 각각 정렬
            r_items.sort(key=lambda x: (int(x[:-1]), x))  # 숫자 오름차순 정렬
            l_items.sort(key=lambda x: (int(x[:-1]), x))  # 숫자 오름차순 정렬

            sorted_items = r_items + l_items
            floor_columns[floor] = sorted_items
    return render_template('racklayout.html', equipment_data=equipment_data, floors=sorted(floors),
                           floor_columns=floor_columns, rack_info_dict=rack_info_dict)


@app.route('/rack_export', methods=['GET', 'POST'])
def rack_export():
    # 현재 테이블 데이터 가져오기
    sql = "SELECT loc1, loc2, servername, charge, maker, model, usize FROM total_asset"
    db = get_db_connection()
    try:
        with db.cursor() as cursor:
            cursor.execute(sql)
            data = cursor.fetchall()
    finally:
        db.close()

    # 컬럼 이름 가져오기
    columns = [column[0] for column in cursor.description]

    # DataFrame 생성
    df = pd.DataFrame(data, columns=columns)

    # loc1의 값이 없는 행 제거
    df = df.dropna(subset=['loc1'])

    selected_floor = request.form.get('floor')
    selected_column = request.form.get('column')
    selected_location = request.form.get('location')


    filtered_df = df[df['loc1'] == f"{selected_floor}-{selected_column}-{selected_location}"]
    filtered_df = filtered_df[['loc2', 'model', 'servername', 'charge', 'usize']]

    # loc2 범위 설정
    loc2_range = range(1, 43)

    # 현재 loc2 값
    existing_loc2 = filtered_df['loc2'].unique()

    # 추가할 데이터프레임 생성
    new_rows = []

    # loc2가 없는 값에 대한 데이터 추가
    for loc2 in loc2_range:
        if loc2 not in existing_loc2:
            # loc2에 해당하는 데이터 추가
            new_row = {'loc2': loc2, 'model': '', 'servername': '', 'charge': '', 'usize': ''}
            new_rows.append(new_row)

    # 새로운 행들을 데이터프레임으로 변환
    new_rows_df = pd.DataFrame(new_rows)

    # 기존 데이터프레임과 새로운 행들을 결합
    filtered_df = pd.concat([filtered_df, new_rows_df], ignore_index=True)

    # loc2 기준으로 내림차순 정렬
    filtered_df.sort_values(by='loc2', ascending=False, inplace=True)
    filtered_df.reset_index(drop=True, inplace=True)

    template_path = 'static/excel/template_rack.xlsx'
    # 엑셀 템플릿 파일 열기
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active  # 기본 시트 선택

    # 데이터 삽입 시작 위치 설정
    start_row = 4
    start_col = 2

    sheet.cell(row=2, column=2).value = f"{selected_floor}-{selected_column}-{selected_location}"

    # 데이터 삽입
    for loc2 in filtered_df['loc2'].unique():  # loc2의 고유 값으로 반복
        rows = filtered_df[filtered_df['loc2'] == loc2]  # loc2에 해당하는 모든 행 선택

        for i, row in rows.iterrows():  # 각 행을 반복
            for j, col in enumerate(filtered_df.columns):  # 열 인덱스 사용
                if j == 0: continue
                cell = sheet.cell(row=start_row + loc2, column=start_col + j)

                # 기존 값이 있을 경우 개행하여 추가
                if cell.value:
                    # cell.value가 int인 경우 문자열로 변환
                    if isinstance(cell.value, int):
                        cell.value = str(cell.value)
                    cell.value += f"\n{row[col]}"  # 여기서 row[col]을 추가
                else:
                    cell.value = row[col]  # 첫 번째 값 설정

    # 셀의 높이를 자동 조정하기 위해 개행 설정
    for row in sheet.iter_rows(min_row=start_row, max_row=start_row + len(filtered_df) - 1, min_col=start_col,
                               max_col=start_col + len(filtered_df.columns) - 1):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    # 변경된 엑셀 파일 저장
    output_path = 'output.xlsx'
    workbook.save(output_path)

    return send_file(output_path, as_attachment=True)


@app.route('/rackview', methods=['GET', 'POST'])
def rackview():
    sql = "SELECT loc1, loc2, servername, charge, maker, model, usize FROM total_asset"
    db = get_db_connection()
    try:
        with db.cursor() as cursor:
            cursor.execute(sql)
            data = cursor.fetchall()
            # 컬럼 이름 가져오기
            columns = [column[0] for column in cursor.description]
            # DataFrame 생성
            df = pd.DataFrame(data, columns=columns)
    finally:
        db.close()

    # loc1의 값이 없는 행 제거
    df = df.dropna(subset=['loc1'])

    # loc 컬럼에서 층, 열, 위치 추출
    floors = set()
    columns_set = set()
    locations = set()

    for loc in df['loc1']:
        try:
            floor, column, location = loc.split('-')
            floors.add(floor)
            columns_set.add(column)
            locations.add(location)
        except ValueError:
            continue

    arg_floor = request.args.get('floor')
    arg_column = request.args.get('column')
    arg_location = request.args.get('location')

    if request.method == 'POST':
        selected_floor = request.form.get('floor')
        selected_column = request.form.get('column')
        selected_location = request.form.get('location')

        # loc 필터링
        filtered_df = df[df['loc1'] == f"{selected_floor}-{selected_column}-{selected_location}"]

        # 결과 저장할 리스트
        result_df = pd.DataFrame()

        # 상단번호 42부터 1까지 반복
        idx = 42
        while idx > 0:
            loc2_rows = filtered_df[filtered_df['loc2'] == idx]
            if not loc2_rows.empty:
                # 중복된 모델명, 서버명, 담당자 수집
                model_names = loc2_rows['model'].dropna().unique().tolist()  # None 값 제거
                server_names = loc2_rows['servername'].dropna().unique().tolist()  # None 값 제거
                charges = loc2_rows['charge'].dropna().unique().tolist()  # None 값 제거
                maker_names = loc2_rows['maker'].dropna().unique().tolist()  # None 값 제거

                # 각 정보 합치기
                maker_combined = ' | '.join(filter(None, maker_names))  # None 값 필터링
                model_combined = ' | '.join(filter(None, model_names))  # None 값 필터링
                server_combined = ' | '.join(filter(None, server_names))  # None 값 필터링
                charge_combined = ' | '.join(filter(None, charges))  # None 값 필터링

                # usize로 병합할 행 수 결정
                usize = int(loc2_rows['usize'].max())
                result_df[idx] = dict(maker=maker_combined, model=model_combined,
                                                  servername=server_combined, charge=charge_combined, usize=usize)
                idx -= usize - 1
            else:
                result_df[idx] = dict(maker='', model='', servername='', charge='', usize=1)
            idx -= 1

        temp_result = []
        for idx in result_df.columns:
            temp_result.append({'loc2': idx, 'maker': result_df[idx].maker, 'model': result_df[idx].model,
                                'servername': result_df[idx].servername, 'charge': result_df[idx].charge,
                                'usize': result_df[idx].usize})

        return jsonify(temp_result)

    # 상면번호 내림차순 정렬
    equipment_data = [{'id': i, 'name': '', 'owner': ''} for i in range(42, 0, -1)]

    # 모든 값이 존재하는지 확인
    if arg_floor and arg_column and arg_location:
        return render_template('rack.html',
                               equipment_data=equipment_data,
                               floors=list(floors),
                               columns=list(columns_set),
                               locations=list(locations),
                               selected_floor=arg_floor,
                               selected_column=arg_column,
                               selected_location=arg_location,
                               auto_fetch=True)

    return render_template('rack.html',
                           equipment_data=equipment_data,
                           floors=list(floors),
                           columns=list(columns_set),
                           locations=list(locations))

@app.route('/get_locations', methods=['GET'])
def get_locations():
    floor = request.args.get('floor')
    column = request.args.get('column')

    sql = "SELECT loc1 FROM total_asset WHERE loc1 LIKE %s"
    db = get_db_connection()
    try:
        with db.cursor() as cursor:
            cursor.execute(sql, (f"{floor}-{column}-%",))  # floor와 column에 따라 필터링
            data = cursor.fetchall()
    finally:
        db.close()

    # loc1에서 위치만 추출
    locations = set(loc['loc1'].split('-')[2] for loc in data)  # loc1에서 위치 추출

    return jsonify(sorted(list(locations)))

@app.route('/get_columns', methods=['GET'])
def get_columns():
    floor = request.args.get('floor')

    sql = "SELECT DISTINCT loc1 FROM total_asset WHERE loc1 LIKE %s"
    db = get_db_connection()
    try:
        with db.cursor() as cursor:
            cursor = db.cursor()
            cursor.execute(sql, (f"{floor}-%",))  # 선택된 floor에 대한 loc1 필터링
            data = cursor.fetchall()
    finally:
        db.close()

    # loc1에서 column만 추출
    columns = set(loc['loc1'].split('-')[1] for loc in data)  # loc1에서 column 추출

    return jsonify(sorted(list(columns)))


class FileForm(FlaskForm):    # 파일 유효성 검사하는 폼 클래스 생성
    files = FileField(validators=[FileRequired('업로드할 파일을 넣어주세요')])

def stamp2real(stamp):
    return datetime.fromtimestamp(stamp)

def info(filename):
    ctime = os.path.getctime(filename) # 만든시간
    mtime = os.path.getmtime(filename) # 수정시간
    size = os.path.getsize(filename) # 파일크기 (단위: bytes)
    return ctime, mtime, size

@app.errorhandler(404) # 404에러 처리
def page_not_found(error):
     return render_template('deny.html', pwd=os.getcwd() + "\\uploads"), 40


@app.route('/fileindex', methods=['GET', 'POST'])
def fileindex():
    form = FileForm()  # 파일 유효성 폼 클래스 인스턴스 생성
    current_path = request.args.get('path', './uploads')

    # 상위 디렉토리 경로 계산
    parent_path = os.path.dirname(current_path)

    if form.validate_on_submit():  # 양식 유효성 검사 + POST인 경우
        f = form.files.data
        f.save(os.path.join(current_path, f.filename))  # 파일 업로드
        return redirect(f'/fileindex?path={current_path}')

    # 새 폴더 생성
    if request.method == 'POST' and 'create_folder' in request.form:
        folder_name = request.form['folder_name']
        os.makedirs(os.path.join(current_path, folder_name), exist_ok=True)
        return redirect(f'/fileindex?path={current_path}')

    # 폴더 삭제 처리
    if request.method == 'POST' and 'delete_folder' in request.form:
        folder_name = request.form['folder_name']
        folder_path = os.path.join(current_path, folder_name)
        if os.path.isdir(folder_path):
            os.rmdir(folder_path)  # 폴더 삭제
            flash(f'폴더 "{folder_name}"가 삭제되었습니다.', 'success')
        else:
            flash(f'폴더 "{folder_name}"가 존재하지 않습니다.', 'error')
        return redirect(f'/fileindex?path={current_path}')

    filelist = os.listdir(current_path)  # 파일 리스트 가져오기
    infos = []
    for name in filelist:
        fileinfo = {}
        full_path = os.path.join(current_path, name)
        ctime = os.path.getctime(full_path)
        mtime = os.path.getmtime(full_path)

        if os.path.isdir(full_path):  # 디렉토리인지 확인
            fileinfo["name"] = name
            fileinfo["create"] = datetime.fromtimestamp(ctime)
            fileinfo["modify"] = datetime.fromtimestamp(mtime)
            fileinfo["size"] = "폴더"
            fileinfo["isfile"] = False
        else:
            size = os.path.getsize(full_path)
            fileinfo["name"] = name
            fileinfo["create"] = datetime.fromtimestamp(ctime)
            fileinfo["modify"] = datetime.fromtimestamp(mtime)
            fileinfo["isfile"] = True
            if size <= 1000000:
                fileinfo["size"] = "%.2f KB" % (size / 1024)
            else:
                fileinfo["size"] = "%.2f MB" % (size / (1024.0 * 1024.0))
        infos.append(fileinfo)  # 각 파일 정보를 모두 리스트로 입력

    return render_template('fileindex.html', form=form,
                           pwd=current_path, parent_path=parent_path, infos=infos)


@app.route('/down/<path:filename>')
def down_page(filename):
    return send_file(os.path.join(app.root_path, 'uploads', filename), as_attachment=True)

@app.route('/del/<path:filename>')
def delete_page(filename):
    os.remove(os.path.join('uploads', filename))  # 파일 삭제
    return redirect('/fileindex')  # 루트 페이지로 이동

@app.route('/export')
def export_asset():
    # 데이터베이스에서 자산 데이터 가져오기
    sql = "SELECT * FROM total_asset"
    db = get_db_connection()
    try:
        with db.cursor() as cursor:
            cursor.execute(sql)
            data = cursor.fetchall()
    finally:
        db.close()

    # 데이터프레임 생성
    df = pd.DataFrame(data)

    # 엑셀 파일로 저장
    today = datetime.now().strftime('%Y%m%d')
    export_filename = f'asset_{today}.xlsx'
    export_filepath = os.path.join(app.root_path, 'exports', export_filename)

    # exports 폴더가 없으면 생성
    if not os.path.exists(os.path.dirname(export_filepath)):
        os.makedirs(os.path.dirname(export_filepath))

    df.to_excel(export_filepath, index=False)

    # 파일을 사용자에게 전송
    return send_file(export_filepath, as_attachment=True)

@app.route('/upload', methods=['POST'])
def upload_file():
    # 업로드 폴더 경로 설정
    app.config['UPLOAD_FOLDER'] = os.path.join(app.root_path, 'uploads')

    # uploads 폴더가 없으면 생성
    if not os.path.exists(app.config['UPLOAD_FOLDER']):
        os.makedirs(app.config['UPLOAD_FOLDER'])

    if 'file' not in request.files:
        return redirect(request.url)

    file = request.files['file']

    if file.filename == '':
        return redirect(request.url)

    if file and file.filename.endswith('.xlsx'):
        # 파일 저장
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filepath)

        # 엑셀 파일 읽기
        df = pd.read_excel(filepath)
        df = df.replace({np.nan: None})

        # 데이터베이스에 저장
        db = get_db_connection()
        try:
            with db.cursor() as cursor:
                # 유효성 검사 결과 저장
                invalid_rows = []

                for index, row in df.iterrows():
                    try:
                        df['isvm'] = df['isvm'].astype(int)
                        df['isoper'] = df['isoper'].astype(int)
                        df['oper'] = df['oper'].astype(int)
                        df['power'] = df['power'].astype(int)
                        df['power'] = df['domain'].astype(int)
                    except (ValueError, TypeError):
                        invalid_rows.append(index + 1)  # 변환 실패 시 유효하지 않은 행으로 간주
                        print(ValueError, TypeError)
                        continue

                # 모든 데이터가 유효할 경우 삽입
                valid_data = []
                for index, row in df.iterrows():
                    valid_data.append((row['itamnum'], row['servername'], row['ip'],
                                       row['hostname'], row['center'], row['loc1'],
                                       row['loc2'], row['isvm'], row['vcenter'],
                                       row['datein'], row['dateout'], row['charge'],
                                       row['charge2'], row['isoper'], row['oper'],
                                       row['power'], row['pdu'], row['os'], row['osver'],
                                       row['maker'], row['model'], row['serial'], row['domain'], ['charge3']))

                sql = """INSERT INTO total_asset (itamnum, servername, ip, hostname, center, loc1, loc2, isvm, vcenter, 
                datein, dateout, charge, charge2, isoper, oper, power, pdu, os, osver, maker, model, serial, domain, 
                charge3) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
                %s)"""
                cursor.executemany(sql, valid_data)
                db.commit()
        finally:
            db.close()

        return redirect(url_for('index'))

    return redirect(request.url)


@app.route('/upload_rv', methods=['POST'])
def upload_rv_file():
    # 업로드 폴더 경로 설정
    app.config['UPLOAD_FOLDER'] = os.path.join(app.root_path, 'uploads')

    # uploads 폴더가 없으면 생성
    if not os.path.exists(app.config['UPLOAD_FOLDER']):
        os.makedirs(app.config['UPLOAD_FOLDER'])

    if 'file' not in request.files:
        return redirect(request.url)

    file = request.files['file']

    if file.filename == '':
        return redirect(request.url)

    if file and file.filename.endswith('.xlsx'):
        # 파일 저장
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filepath)

        # 엑셀 파일 읽기
        df = pd.read_excel(filepath)
        df = df.replace({np.nan: None})

        # 필요한 칼럼으로 필터링하여 새로운 데이터프레임 생성
        filter_df = df[['VM', 'CPUs', 'Memory', 'Primary IP Address', 'Annotation', 'Datacenter', 'Cluster', 'Host',
                        'OS according to the VMware Tools']].copy()

        # 칼럼 이름 변경
        filter_df.columns = ['hostname', 'cpu', 'memory', 'ip', 'servername', 'center', 'Cluster', 'Host', 'os']

        print(filter_df)

        # 데이터베이스에 저장
        db = get_db_connection()
        cursor = db.cursor()

        # 유효성 검사 결과 저장
        invalid_rows = []

        for index, row in filter_df.iterrows():
            try:
                filter_df['cpu'] = filter_df['cpu'].astype(int)
                # filter_df['memory'] = df['memory'].astype(int)
                # filter_df['oper'] = df['oper'].astype(int)
            except (ValueError, TypeError):
                invalid_rows.append(index + 1)  # 변환 실패 시 유효하지 않은 행으로 간주
                print(ValueError, TypeError)
                continue

        # 모든 데이터가 유효할 경우 삽입
        valid_data = []
        for index, row in filter_df.iterrows():
            valid_data.append((row['servername'], row['ip'],
                               row['hostname'], row['center'], 1, row['Cluster'] + " " + row['Host'],
                               1, 1, row['os'], '서버', row['cpu'], row['memory']))

        sql = """INSERT INTO total_asset (servername, ip, hostname, center, isvm, vcenter,
        isoper, oper, os, domain, cpucore, memory) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        cursor.executemany(sql, valid_data)
        db.commit()

        cursor.close()
        db.close()

        return redirect(url_for('index'))

    return redirect(request.url)

@app.route('/index_detail', methods=['GET', 'POST'])
def index_detail():
    selected_columns = request.form.getlist('columns') if request.method == 'POST' else None

    if request.method == 'POST':
        # SQL 쿼리 초기화
        sql = """
            SELECT ta.*, 
                   id.state AS domain_state, 
                   io_isoper.state AS isoper_state, 
                   io_oper.state AS oper_state, 
                   io_power.state AS power_state, 
                   io_os.state AS os_state 
            FROM total_asset ta 
            JOIN info_domain id ON ta.domain = id.domain
            LEFT JOIN info_isoper io_isoper ON ta.isoper = io_isoper.isoper
            LEFT JOIN info_oper io_oper ON ta.oper = io_oper.oper
            LEFT JOIN info_power io_power ON ta.power = io_power.power
            LEFT JOIN info_os io_os ON ta.os = io_os.os
            WHERE 1=1
            ORDER BY ta.dateinsert
        """  # WHERE 절을 항상 true로 시작
        params = []

        # 입력값 가져오기
        itamnum = request.form.get('itamnum', None)
        servername = request.form.get('servername', None)
        ip = request.form.get('ip', None)
        hostname = request.form.get('hostname', None)
        center = request.form.get('center', None)
        loc1 = request.form.get('loc1', None)
        loc2 = request.form.get('loc2', None, type=int)
        isvm = request.form.get('isvm', None, type=int)
        vcenter = request.form.get('vcenter', None, type=int)
        datein = request.form.get('datein', None)
        dateout = request.form.get('dateout', None)
        charge = request.form.get('charge', None)
        charge2 = request.form.get('charge2', None)
        isoper = request.form.get('isoper', None, type=int)
        oper = request.form.get('oper', None, type=int)
        power = request.form.get('power', None, type=int)
        pdu = request.form.get('pdu', None)
        os = request.form.get('os', None)
        osver = request.form.get('osver', None)
        maker = request.form.get('maker', None)
        model = request.form.get('model', None)
        serial = request.form.get('serial', None)
        domain = request.form.get('domain', None)
        charge3 = request.form.get('charge3', None)

        # 조건 추가
        if itamnum:
            sql += " AND itamnum LIKE %s"
            params.append(f'%{itamnum}%')
        if servername:
            sql += " AND servername LIKE %s"
            params.append(f'%{servername}%')
        if ip:
            sql += " AND ip LIKE %s"
            params.append(f'%{ip}%')
        if hostname:
            sql += " AND hostname LIKE %s"
            params.append(f'%{hostname}%')
        if center:
            sql += " AND center LIKE %s"
            params.append(f'%{center}%')
        if loc1:
            sql += " AND loc1 LIKE %s"
            params.append(f'%{loc1}%')
        if loc2 is not None:  # loc2가 0일 경우도 유효한 값으로 처리
            sql += " AND loc2 = %s"
            params.append(loc2)
        if isvm is not None:
            sql += " AND isvm = %s"
            params.append(isvm)
        if vcenter is not None:
            sql += " AND vcenter = %s"
            params.append(vcenter)
        if datein:
            sql += " AND datein LIKE %s"
            params.append(f'%{datein}%')
        if dateout:
            sql += " AND dateout LIKE %s"
            params.append(f'%{dateout}%')
        if charge:
            sql += " AND charge LIKE %s"
            params.append(f'%{charge}%')
        if charge2:
            sql += " AND charge2 LIKE %s"
            params.append(f'%{charge2}%')
        if isoper is not None:
            sql += " AND isoper = %s"
            params.append(isoper)
        if oper is not None:
            sql += " AND oper = %s"
            params.append(oper)
        if power is not None:
            sql += " AND power = %s"
            params.append(power)
        if pdu:
            sql += " AND pdu LIKE %s"
            params.append(f'%{pdu}%')
        if os:
            sql += " AND os LIKE %s"
            params.append(f'%{os}%')
        if osver:
            sql += " AND osver LIKE %s"
            params.append(f'%{osver}%')
        if maker:
            sql += " AND maker LIKE %s"
            params.append(f'%{maker}%')
        if model:
            sql += " AND model LIKE %s"
            params.append(f'%{model}%')
        if serial:
            sql += " AND serial LIKE %s"
            params.append(f'%{serial}%')
        if domain:
            sql += " AND domain LIKE %s"
            params.append(f'%{domain}%')
        if charge3:
            sql += " AND charge3 LIKE %s"
            params.append(f'%{charge3}%')

        # DB에서 데이터 가져오기
        db = get_db_connection()
        try:
            with db.cursor() as cursor:
                cursor.execute(sql, params)
                data = cursor.fetchall()
        finally:
            db.close()

        return render_template('index_detail.html', data=data, selected_columns=selected_columns)

    # GET 요청 시 전체 자산 조회
    sql = """
        SELECT ta.*, 
               id.state AS domain_state, 
               io_isoper.state AS isoper_state, 
               io_oper.state AS oper_state, 
               io_power.state AS power_state, 
               io_os.state AS os_state 
        FROM total_asset ta 
        JOIN info_domain id ON ta.domain = id.domain
        LEFT JOIN info_isoper io_isoper ON ta.isoper = io_isoper.isoper
        LEFT JOIN info_oper io_oper ON ta.oper = io_oper.oper
        LEFT JOIN info_power io_power ON ta.power = io_power.power
        LEFT JOIN info_os io_os ON ta.os = io_os.os
    """
    db = get_db_connection()
    try:
        with db.cursor() as cursor:
            cursor.execute(sql)
            data = cursor.fetchall()
    finally:
        db.close()

    return render_template('index_detail.html', data=data, selected_columns=None)


@app.route('/')
@app.route('/index')
def index():
    # 오늘 날짜와 6개월 전 날짜 계산
    end_date = datetime.now()

    # SQL 쿼리 작성
    sql_graph = """
                SELECT ta.*, 
                       id.state AS domain_state, 
                       io_isoper.state AS isoper_state, 
                       io_oper.state AS oper_state, 
                       io_power.state AS power_state, 
                       io_os.state AS os_state 
                FROM hli_asset.total_asset ta 
                JOIN hli_asset.info_domain id ON ta.domain = id.domain
                LEFT JOIN hli_asset.info_isoper io_isoper ON ta.isoper = io_isoper.isoper
                LEFT JOIN hli_asset.info_oper io_oper ON ta.oper = io_oper.oper
                LEFT JOIN hli_asset.info_power io_power ON ta.power = io_power.power
                LEFT JOIN hli_asset.info_os io_os ON ta.os = io_os.os
                ORDER BY ta.dateinsert
                """

    # 데이터베이스 연결
    db = get_db_connection()
    try:
        with db.cursor() as cursor:
            cursor.execute(sql_graph)
            data_graph = cursor.fetchall()
            columns = [column[0] for column in cursor.description]
            df_graph = pd.DataFrame(data_graph, columns=columns)
    finally:
        db.close()

    # 데이터 변환 및 집계
    df_graph['datein'] = pd.to_datetime(df_graph['datein'])
    df_graph['dateout'] = pd.to_datetime(df_graph['dateout'])

    # 월 단위로 집계
    df_graph['month'] = df_graph['datein'].dt.to_period('M').astype(str)

    # 각 월별 설치 및 폐기 자산 수 집계
    monthly_data = df_graph.groupby('month').agg(
        installed=('os_state', 'size'),  # 설치된 자산 수
        discarded=('dateout', lambda x: x.notnull().sum())  # 폐기된 자산 수
    ).reset_index()

    # 누적 계산
    monthly_data['cumulative_installed'] = monthly_data['installed'].cumsum()
    monthly_data['cumulative_discarded'] = monthly_data['discarded'].cumsum()

    # 총 자산 수 계산
    monthly_data['total_assets'] = monthly_data['cumulative_installed'] - monthly_data['cumulative_discarded']

    # 각 월의 OS 종류 및 state 정보 집계
    df_graph['label'] = df_graph.apply(lambda row: row['os_state'] if row['domain_state'] == '서버' else row['domain_state'], axis=1)
    monthly_counts = df_graph.groupby(['month', 'label']).size().unstack(fill_value=0)

    # 누적 자산 수를 포함한 데이터프레임 생성
    monthly_counts = monthly_counts.reindex(monthly_data['month'], fill_value=0)
    cumulative_counts = monthly_counts.cumsum()

    # x축을 6개로 제한
    cumulative_counts = cumulative_counts.tail(6)
    monthly_data = monthly_data.tail(6)

    fig = px.bar(cumulative_counts, title='전체자산변동', barmode='stack',
                 labels={'value': '개수', 'month': '월'})

    # 누적 총 자산 표시
    fig.add_trace(px.line(monthly_data, x='month', y='total_assets', line_shape='linear').data[0])

    graph = fig.to_html(full_html=False)

    data_card = get_data()

    return render_template('index.html', data_card=data_card, data=data_graph, graph=graph)

@app.route('/write')
def write_asset():
    db = get_db_connection()

    try:
        with db.cursor() as cursor:
            # info_isoper 데이터 가져오기
            cursor.execute("SELECT * FROM info_isoper")
            isoper_options = cursor.fetchall()

            # info_oper 데이터 가져오기
            cursor.execute("SELECT * FROM info_oper")
            oper_options = cursor.fetchall()

            # info_power 데이터 가져오기
            cursor.execute("SELECT * FROM info_power")
            power_options = cursor.fetchall()

            # info_os 데이터 가져오기
            cursor.execute("SELECT * FROM info_os")
            os_options = cursor.fetchall()

            # info_domain 데이터 가져오기
            cursor.execute("SELECT * FROM info_domain")
            domain_options = cursor.fetchall()

    finally:
        db.close()
    print(isoper_options, oper_options, power_options, os_options, domain_options)
    return render_template('write.html',
                           isoper_options=isoper_options,
                           oper_options=oper_options,
                           power_options=power_options,
                           os_options=os_options,
                           domain_options=domain_options)


@app.route('/add', methods=['POST'])
def add_asset():
    sql = """INSERT INTO total_asset (itamnum, servername, ip, hostname, center, loc1, loc2, isvm, vcenter, 
            datein, dateout, charge, charge2, isoper, oper, power, pdu, os, osver, maker, model, serial, domain, 
            charge3) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
            %s)"""

    itamnum = request.form.get('itamnum')
    servername = request.form.get('servername')
    ip = request.form.get('ip')
    hostname = request.form.get('hostname')
    center = request.form.get('center')
    loc1 = request.form.get('loc1')
    loc2 = request.form.get('loc2', type=int)
    isvm = request.form.get('isvm', type=int)
    vcenter = request.form.get('vcenter', type=int)
    datein = request.form.get('datein')
    dateout = request.form.get('dateout')
    charge = request.form.get('charge')
    charge2 = request.form.get('charge2')
    isoper = request.form.get('isoper')
    oper = request.form.get('oper')
    power = request.form.get('power')
    pdu = request.form.get('pdu')
    os = request.form.get('os')
    osver = request.form.get('osver')
    maker = request.form.get('maker')
    model = request.form.get('model')
    serial = request.form.get('serial')
    domain = request.form.get('domain')
    charge3 = request.form.get('charge3')

    # 날짜 형식 검사 및 변환
    try:
        datein = datetime.strptime(datein, '%Y-%m-%d').date()
    except ValueError:
        datein = None  # 오류 발생 시 None으로 설정

    try:
        dateout = datetime.strptime(dateout, '%Y-%m-%d').date()
    except ValueError:
        dateout = None  # 오류 발생 시 None으로 설정

    db = get_db_connection()
    try:
        with db.cursor() as cursor:
            # isoper
            cursor.execute("SELECT isoper FROM hli_asset.info_isoper WHERE state = %s", (isoper,))
            isoper_value = cursor.fetchone()
            isoper = isoper_value['isoper'] if isoper_value else 0  # 딕셔너리에서 'isoper' 키로 값 가져오기

            # oper
            cursor.execute("SELECT oper FROM hli_asset.info_oper WHERE state = %s", (oper,))
            oper_value = cursor.fetchone()
            oper = oper_value['oper'] if oper_value else 0  # 딕셔너리에서 'oper' 키로 값 가져오기

            # power
            cursor.execute("SELECT power FROM hli_asset.info_power WHERE state = %s", (power,))
            power_value = cursor.fetchone()
            power = power_value['power'] if power_value else 0  # 딕셔너리에서 'power' 키로 값 가져오기

            # os
            cursor.execute("SELECT os FROM hli_asset.info_os WHERE state = %s", (os,))
            os_value = cursor.fetchone()
            os = os_value['os'] if os_value else 0  # 딕셔너리에서 'os' 키로 값 가져오기

            # domain
            cursor.execute("SELECT domain FROM hli_asset.info_domain WHERE state = %s", (domain,))
            domain_value = cursor.fetchone()
            domain = domain_value['domain'] if domain_value else 0  # 딕셔너리에서 'domain' 키로 값 가져오기

            # 데이터 삽입
            cursor.execute(sql, (itamnum, servername, ip, hostname, center, loc1, loc2,
                                 isvm, vcenter, datein, dateout, charge, charge2,
                                 isoper, oper, power, pdu, os, osver, maker, model, serial, domain, charge3))
            db.commit()
    finally:
        db.close()
    return redirect(url_for('index'))

@app.route('/edit/<int:pnum>', methods=['GET', 'POST'])
def edit_asset(pnum):
    db = get_db_connection()

    if request.method == 'POST':
        sql = """UPDATE total_asset SET itamnum = %s, servername = %s, ip = %s, 
                      hostname = %s, center = %s, loc1 = %s, loc2 = %s, isvm = %s, 
                      datein = %s, dateout = %s, charge = %s, charge2 = %s, 
                      isoper = %s, oper = %s, power = %s, pdu = %s, os = %s, 
                      osver = %s, maker = %s, model = %s, serial = %s, domain = %s, charge3 = %s 
                      WHERE pnum = %s"""

        # 폼 데이터 가져오기
        itamnum = request.form.get('itamnum')
        servername = request.form.get('servername')
        ip = request.form.get('ip')
        hostname = request.form.get('hostname')
        center = request.form.get('center')
        loc1 = request.form.get('loc1')
        loc2 = request.form.get('loc2', type=int)
        isvm = request.form.get('isvm', type=int)
        datein = request.form.get('datein')
        dateout = request.form.get('dateout')
        charge = request.form.get('charge')
        charge2 = request.form.get('charge2')
        isoper = request.form.get('isoper')
        oper = request.form.get('oper')
        power = request.form.get('power')
        pdu = request.form.get('pdu')
        os = request.form.get('os')
        osver = request.form.get('osver')
        maker = request.form.get('maker')
        model = request.form.get('model')
        serial = request.form.get('serial')
        domain = request.form.get('domain')
        charge3 = request.form.get('charge3')

        # 날짜 형식 검사 및 변환
        try:
            datein = datetime.strptime(datein, '%Y-%m-%d').date()
        except ValueError:
            datein = None  # 오류 발생 시 None으로 설정

        try:
            dateout = datetime.strptime(dateout, '%Y-%m-%d').date()
        except ValueError:
            dateout = None  # 오류 발생 시 None으로 설정

        print(isoper, oper, power, os, domain)

        with db.cursor() as cursor:
            # isoper
            cursor.execute("SELECT isoper FROM info_isoper WHERE state = %s", (isoper,))
            isoper_value = cursor.fetchone()  # fetchone() 사용
            print(isoper_value)
            isoper = isoper_value['isoper'] if isoper_value and len(isoper_value) > 0 else None

            # oper
            cursor.execute("SELECT oper FROM info_oper WHERE state = %s", (oper,))
            oper_value = cursor.fetchone()
            oper = oper_value['oper'] if oper_value and len(oper_value) > 0 else None

            # power
            cursor.execute("SELECT power FROM info_power WHERE state = %s", (power,))
            power_value = cursor.fetchone()
            power = power_value['power'] if power_value and len(power_value) > 0 else None

            # os
            cursor.execute("SELECT os FROM info_os WHERE state = %s", (os,))
            os_value = cursor.fetchone()
            os = os_value['os'] if os_value and len(os_value) > 0 else None

            # domain
            cursor.execute("SELECT domain FROM info_domain WHERE state = %s", (domain,))
            domain_value = cursor.fetchone()
            domain = domain_value['domain'] if domain_value and len(domain_value) > 0 else None
            print(domain)

        try:
            with db.cursor() as cursor:
                cursor.execute(sql, (itamnum, servername, ip, hostname, center, loc1, loc2,
                                     isvm, datein, dateout, charge, charge2,
                                     isoper, oper, power, pdu, os, osver, maker, model, serial, domain, charge3, pnum))
                db.commit()
        finally:
            db.close()

        return redirect(url_for('index_detail'))

    # GET 요청 시 데이터 가져오기
    try:
        with db.cursor() as cursor:
            cursor.execute("""
                SELECT ta.*, 
                       io_isoper.state AS isoper, 
                       io_oper.state AS oper,
                       io_os.state AS os,
                       io_domain.state AS domain,
                       io_power.state AS power
                FROM total_asset ta
                LEFT JOIN info_isoper io_isoper ON ta.isoper = io_isoper.isoper
                LEFT JOIN info_oper io_oper ON ta.oper = io_oper.oper
                LEFT JOIN info_os io_os ON ta.os = io_os.os
                LEFT JOIN info_domain io_domain ON ta.domain = io_domain.domain
                LEFT JOIN info_power io_power ON ta.power = io_power.power
                WHERE ta.pnum = %s
            """, (pnum,))
            data = cursor.fetchone()

            # 옵션 데이터를 가져오기
            cursor.execute("SELECT * FROM info_isoper")
            isoper_options = cursor.fetchall()

            cursor.execute("SELECT * FROM info_oper")
            oper_options = cursor.fetchall()

            cursor.execute("SELECT * FROM info_os")
            os_options = cursor.fetchall()

            cursor.execute("SELECT * FROM info_domain")
            domain_options = cursor.fetchall()

            cursor.execute("SELECT * FROM info_power")
            power_options = cursor.fetchall()

    finally:
        db.close()

    return render_template('edit.html', data=data,
                           isoper_options=isoper_options,
                           oper_options=oper_options,
                           os_options=os_options,
                           domain_options=domain_options,
                           power_options=power_options)


@app.route('/delete/<int:pnum>')
def delete_asset(pnum):
    sql = "DELETE FROM total_asset WHERE pnum = %s"
    db = get_db_connection()
    try:
        with db.cursor() as cursor:
            cursor.execute(sql, (pnum,))
            db.commit()
    finally:
        db.close()
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
