from flask import Blueprint, render_template, redirect, url_for, request, jsonify, flash, send_file, session
import pandas as pd
import numpy as np
from datetime import datetime
import plotly.express as px
from utils.db import execute_query, execute_many, get_db_connection
from utils.cache import cache, invalidate_cache_pattern
import os
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from utils.auto_register import handle_auto_registered_assets
from utils.hierarchy import get_asset_hierarchy

asset_bp = Blueprint('asset', __name__)


@asset_bp.route('/index')
def index():
    """자산 관리 메인 페이지"""
    # 데이터 카드 정보 가져오기
    data_card = get_data()

    # 그래프 데이터 가져오기
    graphs = generate_asset_graph()

    # 자산 데이터 가져오기 (isfix=1 또는 isfix=2)
    sql = """
        SELECT ta.*, 
               id.state AS domain_state, 
               io_isoper.state AS isoper_state, 
               io_oper.state AS oper_state, 
               io_power.state AS power_state, 
               io_os.state AS os_state,
               ig.state AS group_state
        FROM total_asset ta 
        JOIN info_domain id ON ta.domain = id.domain
        LEFT JOIN info_isoper io_isoper ON ta.isoper = io_isoper.isoper
        LEFT JOIN info_oper io_oper ON ta.oper = io_oper.oper
        LEFT JOIN info_power io_power ON ta.power = io_power.power
        LEFT JOIN info_os io_os ON ta.os = io_os.os
        LEFT JOIN info_group ig ON ta.`group` = ig.`group` AND ta.domain = ig.domain
        WHERE ta.isfix IN (1, 2)
        ORDER BY ta.isfix DESC, ta.dateupdate DESC
    """
    data = execute_query(sql)

    return render_template('index.html', data_card=data_card, data=data, graphs=graphs)


@asset_bp.route('/update_isfix', methods=['POST'])
def update_isfix():
    """선택된 자산의 isfix 값을 0으로 업데이트"""
    if request.method == 'POST':
        selected_assets = request.form.getlist('selected_assets[]')

        if selected_assets:
            # 선택된 자산의 isfix 값을 0으로 업데이트
            sql = "UPDATE total_asset SET isfix = 0 WHERE pnum IN (%s)" % ','.join(['%s'] * len(selected_assets))
            execute_query(sql, selected_assets, fetch_all=False)

            flash(f'{len(selected_assets)}개 자산의 정합성 확인이 완료되었습니다.', 'success')
        else:
            flash('선택된 자산이 없습니다.', 'warning')

    return redirect(url_for('asset.index'))


@asset_bp.route('/handle_auto_assets', methods=['POST'])
def handle_auto_assets():
    """자동 등록된 자산 처리"""
    if request.method == 'POST':
        action = request.form.get('action')
        selected_assets = request.form.getlist('selected_assets[]')

        if not action or not selected_assets:
            flash('선택된 자산이 없거나 작업이 지정되지 않았습니다.', 'warning')
            return redirect(url_for('asset.index'))

        success, message = handle_auto_registered_assets(action, selected_assets)

        if success:
            flash(message, 'success')
        else:
            flash(message, 'warning')

    return redirect(url_for('asset.index'))


def get_data():
    """자산 데이터 통계 가져오기"""
    # 데이터베이스 쿼리 - isfix=0이고 사용여부가 '사용'인 자산만 조회
    query = """
        SELECT ta.*, io_isoper.state AS isoper_state 
        FROM total_asset ta
        LEFT JOIN info_isoper io_isoper ON ta.isoper = io_isoper.isoper
        WHERE ta.isfix = 0 AND io_isoper.state = '사용'
    """
    db = get_db_connection()
    try:
        with db.cursor() as cursor:
            cursor.execute(query)
            data = cursor.fetchall()
            # 컬럼 이름 가져오기
            columns = [column[0] for column in cursor.description]
            # DataFrame 생성
            df = pd.DataFrame(data, columns=columns)
    finally:
        db.close()

    # 날짜 형식 변환
    df['datein'] = pd.to_datetime(df['datein'], errors='coerce')

    # 개수 계산
    total_assets = len(df)
    total_servers = len(df[df['domain'] == 0])
    physical_servers = len(df[(df['domain'] == 0) & (df['group'] == 0)])
    virtual_servers = len(df[(df['domain'] == 0) & (df['group'] == 1)])

    # 현재 날짜 기준 계산
    current_date = pd.to_datetime('now')
    current_year = current_date.year
    current_month = current_date.month

    current_year_assets = len(df[pd.to_datetime(df['datein']).dt.year == current_year])
    current_month_assets = len(df[(pd.to_datetime(df['datein']).dt.year == current_year) &
                                  (pd.to_datetime(df['datein']).dt.month == current_month)])

    # 운영구분별 자산 수
    oper_assets = len(df[(df['domain'] == 0) & (df['oper'] == 0)])
    qa_assets = len(df[(df['domain'] == 0) & (df['oper'] == 1)])
    dev_assets = len(df[(df['domain'] == 0) & (df['oper'] == 2)])
    dr_assets = len(df[(df['domain'] == 0) & (df['oper'] == 4)])

    # 센터별 자산 수 (IDC/DR)
    idc_assets = len(df[df['center'].str.contains('IDC', na=False)])
    dr_center_assets = len(df[df['center'].str.contains('DR', na=False)])

    return {
        "total_assets": total_assets,
        "total_servers": total_servers,
        "physical_servers": physical_servers,
        "virtual_servers": virtual_servers,
        "current_year_assets": current_year_assets,
        "current_month_assets": current_month_assets,
        "oper_assets": oper_assets,
        "qa_assets": qa_assets,
        "dev_assets": dev_assets,
        "dr_assets": dr_assets,
        "idc_assets": idc_assets,
        "dr_center_assets": dr_center_assets,
        "current_year": current_year,
        "current_month": current_month
    }


def generate_asset_graph():
    """자산 그래프 생성"""
    # SQL 쿼리 작성 - isfix=0이고 사용여부가 '사용'인 자산만 조회
    sql_graph = """
                SELECT ta.*, 
                       id.state AS domain_state, 
                       io_isoper.state AS isoper_state, 
                       io_oper.state AS oper_state, 
                       io_power.state AS power_state, 
                       io_os.state AS os_state,
                       ig.state AS group_state
                FROM hli_asset.total_asset ta 
                JOIN hli_asset.info_domain id ON ta.domain = id.domain
                LEFT JOIN hli_asset.info_isoper io_isoper ON ta.isoper = io_isoper.isoper
                LEFT JOIN hli_asset.info_oper io_oper ON ta.oper = io_oper.oper
                LEFT JOIN hli_asset.info_power io_power ON ta.power = io_power.power
                LEFT JOIN hli_asset.info_os io_os ON ta.os = io_os.os
                LEFT JOIN hli_asset.info_group ig ON ta.`group` = ig.`group` AND ta.domain = ig.domain
                WHERE ta.isfix = 0 AND io_isoper.state = '사용'
                ORDER BY ta.dateinsert
                """

    # 데이터 가져오기
    db = get_db_connection()
    try:
        with db.cursor() as cursor:
            cursor.execute(sql_graph)
            data_graph = cursor.fetchall()
            columns = [column[0] for column in cursor.description]
            df_graph = pd.DataFrame(data_graph, columns=columns)
    finally:
        db.close()

    # 데이터 변환 및 집계
    df_graph['datein'] = pd.to_datetime(df_graph['datein'])
    df_graph['dateout'] = pd.to_datetime(df_graph['dateout'])

    # 최근 1년 기간 설정
    one_year_ago = pd.Timestamp.now() - pd.DateOffset(years=1)

    # 모든 자산에 월 정보 추가
    df_graph['month'] = df_graph['datein'].dt.to_period('M')

    # 1. 전체 자산의 최근 1년간 누적 그래프
    # 최근 1년간의 월 목록 생성
    recent_months = pd.period_range(start=one_year_ago, end=pd.Timestamp.now(), freq='M')

    # 각 월별, 도메인별 누적 자산 계산
    monthly_data = []

    for month in recent_months:
        # 해당 월 이전에 도입되고 (폐기되지 않았거나 해당 월 이후에 폐기된) 자산 필터링
        month_end = month.to_timestamp(how='end')
        valid_assets = df_graph[(df_graph['datein'] <= month_end) &
                                ((df_graph['dateout'].isna()) | (df_graph['dateout'] > month_end))]

        # 도메인별 자산 수 계산
        domain_counts = valid_assets['domain_state'].value_counts()

        # 결과 저장
        for domain, count in domain_counts.items():
            monthly_data.append({
                'month': month.strftime('%Y-%m'),
                'domain_state': domain,
                'count': count
            })

    # 데이터프레임으로 변환
    monthly_df = pd.DataFrame(monthly_data)

    # 피벗 테이블로 변환
    if not monthly_df.empty:
        monthly_pivot = monthly_df.pivot_table(
            index='month',
            columns='domain_state',
            values='count',
            fill_value=0
        ).reset_index()

        fig1 = px.bar(monthly_pivot, x='month', y=monthly_pivot.columns[1:],
                      title='전체 자산 현황 (최근 1년)',
                      labels={'value': '개수', 'month': '월', 'variable': '도메인'},
                      barmode='stack')
    else:
        # 데이터가 없는 경우 빈 그래프 생성
        fig1 = px.bar(title='전체 자산 현황 (최근 1년)')
        fig1.update_layout(
            xaxis_title='월',
            yaxis_title='개수'
        )

    # 2. 최근 1년간 설치/폐기 막대 그래프
    df_recent = df_graph[df_graph['datein'] >= one_year_ago]
    df_recent['month'] = df_recent['datein'].dt.to_period('M').astype(str)
    monthly_data = df_recent.groupby('month').agg(
        installed=('pnum', 'count'),  # 설치된 자산 수
        discarded=('dateout', lambda x: x.notnull().sum())  # 폐기된 자산 수
    ).reset_index()

    # wide-form에서 long-form으로 변환
    monthly_data_long = pd.melt(
        monthly_data,
        id_vars=['month'],
        value_vars=['installed', 'discarded'],
        var_name='category',
        value_name='count'
    )

    # 카테고리 이름 변경
    category_names = {'installed': '설치', 'discarded': '폐기'}
    monthly_data_long['category'] = monthly_data_long['category'].map(category_names)

    fig2 = px.bar(monthly_data_long, x='month', y='count', color='category',
                  title='최근 1년간 설치/폐기 자산',
                  labels={'count': '개수', 'month': '월', 'category': '구분'},
                  barmode='group',
                  color_discrete_map={'설치': 'green', '폐기': 'red'})
    fig2.update_layout(legend=dict(
        orientation="h",
        yanchor="bottom",
        y=1.02,
        xanchor="right",
        x=1
    ))

    # 3. 도메인별 원형 그래프
    domain_counts = df_graph['domain_state'].value_counts().reset_index()
    domain_counts.columns = ['도메인', '개수']

    fig3 = px.pie(domain_counts, values='개수', names='도메인', title='도메인별 자산 분포')
    fig3.update_traces(textposition='inside', textinfo='percent+label')

    # 4. 서버 도메인의 OS별 원형 그래프
    server_df = df_graph[df_graph['domain_state'] == '서버']
    os_counts = server_df['os_state'].value_counts().reset_index()
    os_counts.columns = ['OS', '개수']

    fig4 = px.pie(os_counts, values='개수', names='OS', title='서버 OS별 분포')
    fig4.update_traces(textposition='inside', textinfo='percent+label')

    # HTML로 그래프 렌더링
    graph1 = fig1.to_html(full_html=False, include_plotlyjs=False)
    graph2 = fig2.to_html(full_html=False, include_plotlyjs=False)
    graph3 = fig3.to_html(full_html=False, include_plotlyjs=False)
    graph4 = fig4.to_html(full_html=False, include_plotlyjs=False)

    return {
        'graph1': graph1,
        'graph2': graph2,
        'graph3': graph3,
        'graph4': graph4
    }


@asset_bp.route('/index_detail', methods=['GET', 'POST'])
def index_detail():
    """자산 상세 검색 페이지"""
    selected_columns = request.form.getlist('columns') if request.method == 'POST' else request.args.getlist('columns')

    # 기본 선택 열 설정 (아무것도 선택되지 않았을 경우)
    if not selected_columns:
        selected_columns = ['domain', 'group', 'servername', 'ip', 'hostname', 'os', 'osver']

    # 항상 포함되어야 하는 필수 열 (pnum은 자세히 링크에 필요)
    required_columns = ['pnum']

    # 열 매핑 정의 (DB 컬럼명 -> 화면 표시명)
    column_mapping = {
        'domain': '도메인',
        'itamnum': 'ITAM자산번호',
        'servername': '서버명',
        'ip': 'IP 주소',
        'hostname': '호스트 이름',
        'center': '센터',
        'loc1': '상면번호',
        'loc2': '상단번호',
        'group': '그룹',
        'vcenter': '상위자산',
        'datein': '설치일자',
        'dateout': '폐기일자',
        'charge': '담당자(정)',
        'charge2': '담당자(부)',
        'isoper': '사용여부',
        'oper': '서비스구분',
        'power': '전원이중화',
        'pdu': 'PDU',
        'os': 'OS',
        'osver': 'OS버전',
        'maker': '제조사',
        'model': '모델명',
        'serial': '시리얼넘버',
        'charge3': '현업담당자',
        'isfix': '정합성 확인 필요',
        'dateupdate': '업데이트 일자'
    }

    # 역매핑 생성 (화면 표시명 -> DB 컬럼명)
    reverse_mapping = {v: k for k, v in column_mapping.items()}

    # SQL 쿼리 초기화 - 선택된 열만 조회
    select_columns = required_columns + selected_columns

    # 중복 제거
    select_columns = list(dict.fromkeys(select_columns))

    # 기본 테이블 열
    base_columns = [f"ta.{col}" for col in select_columns if
                    col not in ['domain', 'isoper', 'oper', 'power', 'os', 'group']]

    # JOIN 테이블 열 (상태 값)
    join_columns = []
    if 'domain' in select_columns:
        join_columns.append("id.state AS domain_state")
    if 'isoper' in select_columns:
        join_columns.append("io_isoper.state AS isoper_state")
    if 'oper' in select_columns:
        join_columns.append("io_oper.state AS oper_state")
    if 'power' in select_columns:
        join_columns.append("io_power.state AS power_state")
    if 'os' in select_columns:
        join_columns.append("io_os.state AS os_state")
    if 'group' in select_columns:
        join_columns.append("ig.state AS group_state")

    # 모든 열 합치기
    all_columns = base_columns + join_columns

    # SQL 쿼리 생성
    sql = f"""
        SELECT {', '.join(all_columns)}
        FROM total_asset ta 
    """

    # JOIN 조건 추가
    if 'domain' in select_columns:
        sql += " JOIN info_domain id ON ta.domain = id.domain"
    if 'isoper' in select_columns:
        sql += " LEFT JOIN info_isoper io_isoper ON ta.isoper = io_isoper.isoper"
    if 'oper' in select_columns:
        sql += " LEFT JOIN info_oper io_oper ON ta.oper = io_oper.oper"
    if 'power' in select_columns:
        sql += " LEFT JOIN info_power io_power ON ta.power = io_power.power"
    if 'os' in select_columns:
        sql += " LEFT JOIN info_os io_os ON ta.os = io_os.os"
    if 'group' in select_columns:
        sql += " LEFT JOIN info_group ig ON ta.`group` = ig.`group` AND ta.domain = ig.domain"

    sql += " WHERE 1=1"
    params = []

    # 기본적으로 isfix=0인 자산만 검색
    if 'isfix' not in request.form and 'isfix' not in request.args:
        sql += " AND ta.isfix = 0"

    # 계층 구조 필터링
    domain_filter = request.args.get('domain', type=int)
    group_filter = request.args.get('group', type=int)

    if domain_filter is not None:
        sql += " AND ta.domain = %s"
        params.append(domain_filter)

    if group_filter is not None:
        sql += " AND ta.`group` = %s"
        params.append(group_filter)

    if request.method == 'POST':
        # 입력값 가져오기
        itamnum = request.form.get('itamnum', None)
        servername = request.form.get('servername', None)
        ip = request.form.get('ip', None)
        hostname = request.form.get('hostname', None)
        center = request.form.get('center', None)
        loc1 = request.form.get('loc1', None)
        loc2 = request.form.get('loc2', None, type=int)
        group = request.form.get('group', None, type=int)
        vcenter = request.form.get('vcenter', None, type=int)
        datein = request.form.get('datein', None)
        dateout = request.form.get('dateout', None)
        charge = request.form.get('charge', None)
        charge2 = request.form.get('charge2', None)
        isoper = request.form.get('isoper', None, type=int)
        oper = request.form.get('oper', None, type=int)
        power = request.form.get('power', None, type=int)
        pdu = request.form.get('pdu', None)
        os = request.form.get('os', None)
        osver = request.form.get('osver', None)
        maker = request.form.get('maker', None)
        model = request.form.get('model', None)
        serial = request.form.get('serial', None)
        domain = request.form.get('domain', None)
        charge3 = request.form.get('charge3', None)
        isfix = request.form.get('isfix', None, type=int)

        # 조건 추가
        if itamnum:
            sql += " AND ta.itamnum LIKE %s"
            params.append(f'%{itamnum}%')
        if servername:
            sql += " AND ta.servername LIKE %s"
            params.append(f'%{servername}%')
        if ip:
            sql += " AND ta.ip LIKE %s"
            params.append(f'%{ip}%')
        if hostname:
            sql += " AND ta.hostname LIKE %s"
            params.append(f'%{hostname}%')
        if center:
            sql += " AND ta.center LIKE %s"
            params.append(f'%{center}%')
        if loc1:
            sql += " AND ta.loc1 LIKE %s"
            params.append(f'%{loc1}%')
        if loc2 is not None:
            sql += " AND ta.loc2 = %s"
            params.append(loc2)
        if group is not None:
            sql += " AND ta.`group` = %s"
            params.append(group)
        if vcenter is not None:
            sql += " AND ta.vcenter = %s"
            params.append(vcenter)
        if datein:
            sql += " AND ta.datein LIKE %s"
            params.append(f'%{datein}%')
        if dateout:
            sql += " AND ta.dateout LIKE %s"
            params.append(f'%{dateout}%')
        if charge:
            sql += " AND ta.charge LIKE %s"
            params.append(f'%{charge}%')
        if charge2:
            sql += " AND ta.charge2 LIKE %s"
            params.append(f'%{charge2}%')
        if isoper is not None:
            sql += " AND ta.isoper = %s"
            params.append(isoper)
        if oper is not None:
            sql += " AND ta.oper = %s"
            params.append(oper)
        if power is not None:
            sql += " AND ta.power = %s"
            params.append(power)
        if pdu:
            sql += " AND ta.pdu LIKE %s"
            params.append(f'%{pdu}%')
        if os:
            sql += " AND ta.os LIKE %s"
            params.append(f'%{os}%')
        if osver:
            sql += " AND ta.osver LIKE %s"
            params.append(f'%{osver}%')
        if maker:
            sql += " AND ta.maker LIKE %s"
            params.append(f'%{maker}%')
        if model:
            sql += " AND ta.model LIKE %s"
            params.append(f'%{model}%')
        if serial:
            sql += " AND ta.serial LIKE %s"
            params.append(f'%{serial}%')
        if domain:
            sql += " AND ta.domain = %s"
            params.append(domain)
        if charge3:
            sql += " AND ta.charge3 LIKE %s"
            params.append(f'%{charge3}%')
        if isfix is not None:
            sql += " AND ta.isfix = %s"
            params.append(isfix)

    # 정렬 추가
    sql += " ORDER BY ta.dateupdate DESC LIMIT 1000"  # 성능을 위해 결과 제한

    # 마지막 검색 쿼리와 파라미터를 세션에 저장 (내보내기용)
    session['last_search_query'] = sql
    session['last_search_params'] = params

    # 데이터 가져오기
    data = execute_query(sql, params)

    # 계층 구조 가져오기
    hierarchy = get_asset_hierarchy(domain=domain_filter, group=group_filter)

    # 도메인 및 그룹 옵션 가져오기
    domain_options = execute_query("SELECT * FROM info_domain")

    group_options = []
    if domain_filter is not None:
        group_sql = "SELECT * FROM info_group WHERE domain = %s"
        group_options = execute_query(group_sql, (domain_filter,))

    return render_template('index_detail.html',
                           data=data,
                           selected_columns=selected_columns,
                           column_mapping=column_mapping,
                           hierarchy=hierarchy,
                           domain_options=domain_options,
                           group_options=group_options,
                           domain_filter=domain_filter,
                           group_filter=group_filter)


@asset_bp.route('/get_asset_details/<int:pnum>')
def get_asset_details(pnum):
    """자산 상세 정보 가져오기 (AJAX 요청용)"""
    sql = """
        SELECT ta.*, 
               id.state AS domain_state, 
               io_isoper.state AS isoper_state, 
               io_oper.state AS oper_state, 
               io_power.state AS power_state, 
               io_os.state AS os_state,
               ig.state AS group_state
        FROM total_asset ta 
        JOIN info_domain id ON ta.domain = id.domain
        LEFT JOIN info_isoper io_isoper ON ta.isoper = io_isoper.isoper
        LEFT JOIN info_oper io_oper ON ta.oper = io_oper.oper
        LEFT JOIN info_power io_power ON ta.power = io_power.power
        LEFT JOIN info_os io_os ON ta.os = io_os.os
        LEFT JOIN info_group ig ON ta.`group` = ig.`group` AND ta.domain = ig.domain
        WHERE ta.pnum = %s
    """

    data = execute_query(sql, (pnum,), fetch_all=False)

    # vcenter 값이 있는 경우 상위 자산 정보 가져오기
    if data and data.get('group') == 1 and data.get('vcenter'):
        parent_sql = """
            SELECT pnum, servername, hostname, ip
            FROM total_asset
            WHERE pnum = %s
        """
        parent_data = execute_query(parent_sql, (data['vcenter'],), fetch_all=False)
        if parent_data:
            data['parent_asset'] = parent_data

    return jsonify(data)


@asset_bp.route('/get_groups', methods=['GET'])
def get_groups():
    """도메인에 따른 그룹 정보 가져오기"""
    domain = request.args.get('domain', type=int)

    if domain is None:
        return jsonify([])

    sql = "SELECT `group`, state FROM info_group WHERE domain = %s"
    groups = execute_query(sql, (domain,))

    return jsonify(groups)


@asset_bp.route('/search_assets', methods=['GET'])
def search_assets():
    """자산 검색 API (AJAX 요청용)"""
    search_term = request.args.get('term', '')
    group = request.args.get('group', 0, type=int)  # 기본값은 물리 서버(0)

    # 검색어가 없으면 빈 결과 반환
    if not search_term:
        return jsonify([])

    # 그룹 필터링 조건 추가
    sql = """
        SELECT pnum, servername, hostname, ip
        FROM total_asset
        WHERE `group` = %s AND (servername LIKE %s OR hostname LIKE %s OR ip LIKE %s)
        LIMIT 20
    """

    search_param = f'%{search_term}%'
    results = execute_query(sql, (group, search_param, search_param, search_param))

    return jsonify(results)


@asset_bp.route('/write')
def write_asset():
    """자산 추가 페이지"""
    # 옵션 데이터 가져오기
    isoper_options = execute_query("SELECT * FROM info_isoper")
    oper_options = execute_query("SELECT * FROM info_oper")
    power_options = execute_query("SELECT * FROM info_power")
    os_options = execute_query("SELECT * FROM info_os")
    domain_options = execute_query("SELECT * FROM info_domain")

    # 계층 구조 가져오기
    hierarchy = get_asset_hierarchy()

    return render_template('write.html',
                           isoper_options=isoper_options,
                           oper_options=oper_options,
                           power_options=power_options,
                           os_options=os_options,
                           domain_options=domain_options,
                           hierarchy=hierarchy)


@asset_bp.route('/add', methods=['POST'])
def add_asset():
    """자산 추가 처리"""
    sql = """INSERT INTO total_asset (itamnum, servername, ip, hostname, center, loc1, loc2, `group`, vcenter, 
            datein, dateout, charge, charge2, isoper, oper, power, pdu, os, osver, maker, model, serial, domain, 
            charge3, usize, cpucore, memory, isfix, dateupdate) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
            %s, %s, %s, %s, %s, %s)"""

    # 폼 데이터 가져오기
    itamnum = request.form.get('itamnum')
    servername = request.form.get('servername')
    ip = request.form.get('ip')
    hostname = request.form.get('hostname')
    center = request.form.get('center')
    loc1 = request.form.get('loc1')
    loc2 = request.form.get('loc2', type=int)
    group = request.form.get('group', type=int)

    # vcenter 값 처리 - group이 1(논리)인 경우에만 vcenter 값을 사용
    vcenter = None
    if group == 1:
        vcenter = request.form.get('vcenter', type=int)
        # vcenter 값이 없거나 0인 경우 None으로 설정
        if not vcenter:
            vcenter = None

    datein = request.form.get('datein')
    dateout = request.form.get('dateout')
    charge = request.form.get('charge')
    charge2 = request.form.get('charge2')
    isoper_state = request.form.get('isoper')
    oper_state = request.form.get('oper')
    power_state = request.form.get('power')
    pdu = request.form.get('pdu')
    os_state = request.form.get('os')
    osver = request.form.get('osver')
    maker = request.form.get('maker')
    model = request.form.get('model')
    serial = request.form.get('serial')
    domain_state = request.form.get('domain')
    charge3 = request.form.get('charge3')
    usize = request.form.get('usize', type=int, default=1)
    cpucore = request.form.get('cpucore', type=int)
    memory = request.form.get('memory', type=int)
    isfix = 1  # 새로 추가된 자산은 기본적으로 정합성 확인이 필요
    dateupdate = datetime.now()  # 현재 시간으로 업데이트 일자 설정

    # 날짜 형식 검사 및 변환
    try:
        datein = datetime.strptime(datein, '%Y-%m-%d').date() if datein else None
    except ValueError:
        datein = None

    try:
        dateout = datetime.strptime(dateout, '%Y-%m-%d').date() if dateout else None
    except ValueError:
        dateout = None

    # 코드 값 조회
    db = get_db_connection()
    try:
        with db.cursor() as cursor:
            # isoper
            cursor.execute("SELECT isoper FROM info_isoper WHERE state = %s", (isoper_state,))
            isoper_value = cursor.fetchone()
            isoper = isoper_value['isoper'] if isoper_value else 0

            # oper
            cursor.execute("SELECT oper FROM info_oper WHERE state = %s", (oper_state,))
            oper_value = cursor.fetchone()
            oper = oper_value['oper'] if oper_value else 0

            # power
            cursor.execute("SELECT power FROM info_power WHERE state = %s", (power_state,))
            power_value = cursor.fetchone()
            power = power_value['power'] if power_value else 0

            # os
            cursor.execute("SELECT os FROM info_os WHERE state = %s", (os_state,))
            os_value = cursor.fetchone()
            os_code = os_value['os'] if os_value else 0

            # domain
            cursor.execute("SELECT domain FROM info_domain WHERE state = %s", (domain_state,))
            domain_value = cursor.fetchone()
            domain = domain_value['domain'] if domain_value else 0

            # 데이터 삽입
            cursor.execute(sql, (itamnum, servername, ip, hostname, center, loc1, loc2,
                                 group, vcenter, datein, dateout, charge, charge2,
                                 isoper, oper, power, pdu, os_code, osver, maker, model, serial, domain, charge3,
                                 usize, cpucore, memory, isfix, dateupdate))
            db.commit()
    finally:
        db.close()

    # 캐시 무효화
    invalidate_cache_pattern('index_data')
    invalidate_cache_pattern('asset_data')
    invalidate_cache_pattern('asset_graph')

    return redirect(url_for('asset.index'))


@asset_bp.route('/edit/<int:pnum>', methods=['GET', 'POST'])
def edit_asset(pnum):
    """자산 수정 페이지 및 처리"""
    if request.method == 'POST':
        sql = """UPDATE total_asset SET itamnum = %s, servername = %s, ip = %s, 
                      hostname = %s, center = %s, loc1 = %s, loc2 = %s, `group` = %s, vcenter = %s,
                      datein = %s, dateout = %s, charge = %s, charge2 = %s, 
                      isoper = %s, oper = %s, power = %s, pdu = %s, os = %s, 
                      osver = %s, maker = %s, model = %s, serial = %s, domain = %s, charge3 = %s,
                      usize = %s, cpucore = %s, memory = %s, isfix = %s, dateupdate = %s
                      WHERE pnum = %s"""

        # 폼 데이터 가져오기
        itamnum = request.form.get('itamnum')
        servername = request.form.get('servername')
        ip = request.form.get('ip')
        hostname = request.form.get('hostname')
        center = request.form.get('center')
        loc1 = request.form.get('loc1')
        loc2 = request.form.get('loc2', type=int)
        group = request.form.get('group', type=int)

        # vcenter 값 처리 - group이 1(논리)인 경우에만 vcenter 값을 사용
        vcenter = None
        if group == 1:
            vcenter = request.form.get('vcenter', type=int)
            # vcenter 값이 없거나 0인 경우 None으로 설정
            if not vcenter:
                vcenter = None

        datein = request.form.get('datein')
        dateout = request.form.get('dateout')
        charge = request.form.get('charge')
        charge2 = request.form.get('charge2')
        isoper_state = request.form.get('isoper')
        oper_state = request.form.get('oper')
        power_state = request.form.get('power')
        pdu = request.form.get('pdu')
        os_state = request.form.get('os')
        osver = request.form.get('osver')
        maker = request.form.get('maker')
        model = request.form.get('model')
        serial = request.form.get('serial')
        domain_state = request.form.get('domain')
        charge3 = request.form.get('charge3')
        usize = request.form.get('usize', type=int, default=1)
        cpucore = request.form.get('cpucore', type=int)
        memory = request.form.get('memory', type=int)
        isfix = 1  # 수정된 자산은 기본적으로 정합성 확인이 필요
        dateupdate = datetime.now()  # 현재 시간으로 업데이트 일자 설정

        # 날짜 형식 검사 및 변환
        try:
            datein = datetime.strptime(datein, '%Y-%m-%d').date() if datein else None
        except ValueError:
            datein = None

        try:
            dateout = datetime.strptime(dateout, '%Y-%m-%d').date() if dateout else None
        except ValueError:
            dateout = None

        # 코드 값 조회
        db = get_db_connection()
        try:
            with db.cursor() as cursor:
                # isoper
                cursor.execute("SELECT isoper FROM info_isoper WHERE state = %s", (isoper_state,))
                isoper_value = cursor.fetchone()
                isoper = isoper_value['isoper'] if isoper_value else None

                # oper
                cursor.execute("SELECT oper FROM info_oper WHERE state = %s", (oper_state,))
                oper_value = cursor.fetchone()
                oper = oper_value['oper'] if oper_value else None

                # power
                cursor.execute("SELECT power FROM info_power WHERE state = %s", (power_state,))
                power_value = cursor.fetchone()
                power = power_value['power'] if power_value else None

                # os
                cursor.execute("SELECT os FROM info_os WHERE state = %s", (os_state,))
                os_value = cursor.fetchone()
                os_code = os_value['os'] if os_value else None

                # domain
                cursor.execute("SELECT domain FROM info_domain WHERE state = %s", (domain_state,))
                domain_value = cursor.fetchone()
                domain = domain_value['domain'] if domain_value else None

                # 데이터 업데이트
                cursor.execute(sql, (itamnum, servername, ip, hostname, center, loc1, loc2,
                                     group, vcenter, datein, dateout, charge, charge2,
                                     isoper, oper, power, pdu, os_code, osver, maker, model, serial, domain, charge3,
                                     usize, cpucore, memory, isfix, dateupdate, pnum))
                db.commit()
        finally:
            db.close()

        # 캐시 무효화
        invalidate_cache_pattern('index_data')
        invalidate_cache_pattern('asset_data')
        invalidate_cache_pattern('asset_graph')

        return redirect(url_for('asset.index_detail'))

    # GET 요청 시 데이터 가져오기
    sql = """
        SELECT ta.*, 
               io_isoper.state AS isoper_state, 
               io_oper.state AS oper_state,
               io_os.state AS os_state,
               io_domain.state AS domain_state,
               io_power.state AS power_state,
               ig.state AS group_state
        FROM total_asset ta
        LEFT JOIN info_isoper io_isoper ON ta.isoper = io_isoper.isoper
        LEFT JOIN info_oper io_oper ON ta.oper = io_oper.oper
        LEFT JOIN info_os io_os ON ta.os = io_os.os
        LEFT JOIN info_domain io_domain ON ta.domain = io_domain.domain
        LEFT JOIN info_power io_power ON ta.power = io_power.power
        LEFT JOIN info_group ig ON ta.`group` = ig.`group` AND ta.domain = ig.domain
        WHERE ta.pnum = %s
    """
    data = execute_query(sql, (pnum,), fetch_all=False)

    # 상위 자산 정보 가져오기 (vcenter가 있는 경우)
    parent_asset = None
    if data and data.get('group') == 1 and data.get('vcenter'):
        parent_sql = """
            SELECT pnum, servername, hostname, ip
            FROM total_asset
            WHERE pnum = %s
        """
        parent_asset = execute_query(parent_sql, (data['vcenter'],), fetch_all=False)

    # 하위 자산 정보 가져오기 (현재 자산의 pnum을 vcenter로 가지고 있는 자산들)
    child_assets = []
    child_sql = """
        SELECT pnum, servername, hostname, ip
        FROM total_asset
        WHERE vcenter = %s AND `group` = 1
        ORDER BY servername
    """
    child_assets = execute_query(child_sql, (pnum,))

    # 옵션 데이터 가져오기
    isoper_options = execute_query("SELECT * FROM info_isoper")
    oper_options = execute_query("SELECT * FROM info_oper")
    os_options = execute_query("SELECT * FROM info_os")
    domain_options = execute_query("SELECT * FROM info_domain")
    power_options = execute_query("SELECT * FROM info_power")

    # 도메인에 따른 그룹 옵션 가져오기
    group_options = []
    if data and data.get('domain') is not None:
        group_sql = "SELECT * FROM info_group WHERE domain = %s"
        group_options = execute_query(group_sql, (data['domain'],))

    # 계층 구조 가져오기
    hierarchy = get_asset_hierarchy(pnum=pnum, domain=data.get('domain'), group=data.get('group'))

    return render_template('edit.html',
                           data=data,
                           parent_asset=parent_asset,
                           child_assets=child_assets,
                           isoper_options=isoper_options,
                           oper_options=oper_options,
                           os_options=os_options,
                           domain_options=domain_options,
                           power_options=power_options,
                           group_options=group_options,
                           hierarchy=hierarchy)


@asset_bp.route('/delete/<int:pnum>')
def delete_asset(pnum):
    """자산 삭제 처리"""
    sql = "DELETE FROM total_asset WHERE pnum = %s"
    execute_query(sql, (pnum,), fetch_all=False)

    # 캐시 무효화
    invalidate_cache_pattern('index_data')
    invalidate_cache_pattern('asset_data')
    invalidate_cache_pattern('asset_graph')

    return redirect(url_for('asset.index'))


@asset_bp.route('/export')
def export_asset():
    """자산 데이터 내보내기"""
    # 데이터베이스에서 자산 데이터 가져오기 - JOIN을 사용하여 코드값 대신 설명 가져오기
    sql = """
        SELECT 
            ta.pnum, ta.itamnum, ta.servername, ta.ip, ta.hostname, ta.center, 
            ta.loc1, ta.loc2, ig.state AS group_state, ta.vcenter, ta.datein, ta.dateout, 
            ta.charge, ta.charge2, io_isoper.state AS isoper, io_oper.state AS oper, 
            io_power.state AS power, ta.pdu, io_os.state AS os, ta.osver, 
            ta.maker, ta.model, ta.serial, io_domain.state AS domain, ta.charge3,
            ta.usize, ta.vmpnum, ta.dateinsert, ta.cpucore, ta.memory, ta.dateupdate, ta.isfix
        FROM total_asset ta
        LEFT JOIN info_isoper io_isoper ON ta.isoper = io_isoper.isoper
        LEFT JOIN info_oper io_oper ON ta.oper = io_oper.oper
        LEFT JOIN info_power io_power ON ta.power = io_power.power
        LEFT JOIN info_os io_os ON ta.os = io_os.os
        LEFT JOIN info_domain io_domain ON ta.domain = io_domain.domain
        LEFT JOIN info_group ig ON ta.`group` = ig.`group` AND ta.domain = ig.domain
    """
    data = execute_query(sql)

    # 데이터프레임 생성
    df = pd.DataFrame(data)

    # 컬럼명 한글로 변경
    column_mapping = {
        'pnum': '자산고유번호',
        'itamnum': 'ITAM자산번호',
        'servername': '서버명',
        'ip': 'IP 주소',
        'hostname': '호스트 이름',
        'center': '센터',
        'loc1': '상면번호',
        'loc2': '상단번호',
        'group_state': '그룹',
        'vcenter': '상위 자산',
        'datein': '설치일자',
        'dateout': '폐기일자',
        'charge': '담당자(정)',
        'charge2': '담당자(부)',
        'isoper': '사용여부',
        'oper': '서비스구분',
        'power': '전원이중화',
        'pdu': 'PDU',
        'os': 'OS',
        'osver': 'OS버전',
        'maker': '제조사',
        'model': '모델',
        'serial': '시리얼넘버',
        'domain': '도메인',
        'charge3': '현업담당자',
        'usize': '장비크기',
        'vmpnum': 'VM번호',
        'dateinsert': '등록일시',
        'cpucore': '물리코어',
        'memory': '메모리크기',
        'dateupdate': '업데이트일시',
        'isfix': '변경확인'
    }

    # 컬럼명 변경
    df.rename(columns=column_mapping, inplace=True)

    # 변경확인 값 변환 (0 -> '확인완료', 1 -> '확인필요')
    if '변경확인' in df.columns:
        df['변경확인'] = df['변경확인'].map({0: '확인완료', 1: '확인필요'})

    # 엑셀 파일로 저장
    today = datetime.now().strftime('%Y%m%d')
    export_filename = f'asset_{today}.xlsx'
    export_filepath = os.path.join(os.getcwd(), 'exports', export_filename)

    # exports 폴더가 없으면 생성
    if not os.path.exists(os.path.dirname(export_filepath)):
        os.makedirs(os.path.dirname(export_filepath))

    # 엑셀 파일 생성
    with pd.ExcelWriter(export_filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='자산목록')

        # 열 너비 자동 조정
        worksheet = writer.sheets['자산목록']
        for i, column in enumerate(df.columns):
            column_width = max(df[column].astype(str).map(len).max(), len(column) + 2)
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = column_width

    return send_file(export_filepath, as_attachment=True)


@asset_bp.route('/export_filtered_asset')
def export_filtered_asset():
    """현재 검색된 자산 데이터만 내보내기"""
    # 세션에서 마지막 검색 쿼리와 파라미터 가져오기
    last_search_query = session.get('last_search_query')
    last_search_params = session.get('last_search_params', [])

    if not last_search_query:
        # 검색 쿼리가 없으면 기본 쿼리 사용 (isfix=0인 자산만)
        last_search_query = """
            SELECT 
                ta.pnum, ta.itamnum, ta.servername, ta.ip, ta.hostname, ta.center, 
                ta.loc1, ta.loc2, ig.state AS group_state, ta.vcenter, ta.datein, ta.dateout, 
                ta.charge, ta.charge2, io_isoper.state AS isoper, io_oper.state AS oper, 
                io_power.state AS power, ta.pdu, io_os.state AS os, ta.osver, 
                ta.maker, ta.model, ta.serial, io_domain.state AS domain, ta.charge3,
                ta.usize, ta.vmpnum, ta.dateinsert, ta.cpucore, ta.memory, ta.dateupdate, ta.isfix
            FROM total_asset ta
            LEFT JOIN info_isoper io_isoper ON ta.isoper = io_isoper.isoper
            LEFT JOIN info_oper io_oper ON ta.oper = io_oper.oper
            LEFT JOIN info_power io_power ON ta.power = io_power.power
            LEFT JOIN info_os io_os ON ta.os = io_os.os
            LEFT JOIN info_domain io_domain ON ta.domain = io_domain.domain
            LEFT JOIN info_group ig ON ta.`group` = ig.`group` AND ta.domain = ig.domain
            WHERE ta.isfix = 0
        """
        last_search_params = []

    # 데이터 가져오기
    data = execute_query(last_search_query, last_search_params)

    # 데이터프레임 생성
    df = pd.DataFrame(data)

    # 컬럼명 한글로 변경
    column_mapping = {
        'pnum': '자산고유번호',
        'itamnum': 'ITAM자산번호',
        'servername': '서버명',
        'ip': 'IP 주소',
        'hostname': '호스트 이름',
        'center': '센터',
        'loc1': '상면번호',
        'loc2': '상단번호',
        'group_state': '그룹',
        'vcenter': '상위 자산',
        'datein': '설치일자',
        'dateout': '폐기일자',
        'charge': '담당자(정)',
        'charge2': '담당자(부)',
        'isoper': '사용여부',
        'oper': '서비스구분',
        'power': '전원이중화',
        'pdu': 'PDU',
        'os': 'OS',
        'osver': 'OS버전',
        'maker': '제조사',
        'model': '모델',
        'serial': '시리얼넘버',
        'domain': '도메인',
        'charge3': '현업담당자',
        'usize': '장비크기',
        'vmpnum': 'VM번호',
        'dateinsert': '등록일시',
        'cpucore': '물리코어',
        'memory': '메모리크기',
        'dateupdate': '업데이트일시',
        'isfix': '변경확인'
    }

    # 컬럼명 변경
    df.rename(columns=column_mapping, inplace=True)

    # 변경확인 값 변환 (0 -> '확인완료', 1 -> '확인필요')
    if '변경확인' in df.columns:
        df['변경확인'] = df['변경확인'].map({0: '확인완료', 1: '확인필요'})

    # 엑셀 파일로 저장
    today = datetime.now().strftime('%Y%m%d')
    export_filename = f'asset_filtered_{today}.xlsx'
    export_filepath = os.path.join(os.getcwd(), 'exports', export_filename)

    # exports 폴더가 없으면 생성
    if not os.path.exists(os.path.dirname(export_filepath)):
        os.makedirs(os.path.dirname(export_filepath))

    # 엑셀 파일 생성
    with pd.ExcelWriter(export_filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='검색결과')

        # 열 너비 자동 조정
        worksheet = writer.sheets['검색결과']
        for i, column in enumerate(df.columns):
            column_width = max(df[column].astype(str).map(len).max(), len(column) + 2)
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = column_width

    return send_file(export_filepath, as_attachment=True)


@asset_bp.route('/download_template')
def download_template():
    """자산 일괄 등록을 위한 엑셀 템플릿 다운로드"""
    # 템플릿 파일 경로
    template_path = os.path.join(os.getcwd(), 'exports', 'asset_template.xlsx')

    # 템플릿 파일 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "자산 등록"

    # 스타일 정의
    header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 헤더 행 추가
    headers = [
        '서버명*', 'IP 주소*', '호스트 이름', 'ITAM자산번호', '센터', '상면번호', '상단번호',
        '도메인*', '그룹*', '설치일자', '폐기일자', '담당자(정)', '담당자(부)', '사용여부*',
        '서비스구분*', '전원이중화', 'PDU', 'OS*', 'OS버전', '제조사', '모델',
        '시리얼넘버', '현업담당자', '장비크기(U)', '물리코어', '메모리(GB)', '상위자산'
    ]

    # 필수 입력 필드 표시
    required_fields = ['서버명*', 'IP 주소*', '도메인*', '그룹*', '사용여부*', '서비스구분*', 'OS*']

    # 헤더 행 스타일 적용
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_alignment

    # 열 너비 설정
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) * 1.5, 15)

    # 데이터 유효성 검사 설정
    # 도메인 드롭다운
    domain_options = execute_query("SELECT state FROM info_domain")
    domain_list = ','.join([f'"{option["state"]}"' for option in domain_options])
    domain_validation = DataValidation(type="list", formula1=f'"{domain_list}"', allow_blank=True)
    domain_validation.error = "올바른 도메인을 선택하세요."
    domain_validation.errorTitle = "입력 오류"
    ws.add_data_validation(domain_validation)
    domain_validation.add(f"H2:H1000")

    # 그룹 드롭다운 (도메인별로 다르므로 일반적인 목록만 제공)
    group_options = execute_query("SELECT DISTINCT state FROM info_group")
    group_list = ','.join([f'"{option["state"]}"' for option in group_options])
    group_validation = DataValidation(type="list", formula1=f'"{group_list}"', allow_blank=True)
    group_validation.error = "올바른 그룹을 선택하세요."
    group_validation.errorTitle = "입력 오류"
    ws.add_data_validation(group_validation)
    group_validation.add(f"I2:I1000")

    # 사용여부 드롭다운
    isoper_options = execute_query("SELECT state FROM info_isoper")
    isoper_list = ','.join([f'"{option["state"]}"' for option in isoper_options])
    isoper_validation = DataValidation(type="list", formula1=f'"{isoper_list}"', allow_blank=True)
    isoper_validation.error = "올바른 사용여부를 선택하세요."
    isoper_validation.errorTitle = "입력 오류"
    ws.add_data_validation(isoper_validation)
    isoper_validation.add(f"N2:N1000")

    # 서비스구분 드롭다운
    oper_options = execute_query("SELECT state FROM info_oper")
    oper_list = ','.join([f'"{option["state"]}"' for option in oper_options])
    oper_validation = DataValidation(type="list", formula1=f'"{oper_list}"', allow_blank=True)
    oper_validation.error = "올바른 서비스구분을 선택하세요."
    oper_validation.errorTitle = "입력 오류"
    ws.add_data_validation(oper_validation)
    oper_validation.add(f"O2:O1000")

    # 전원이중화 드롭다운
    power_options = execute_query("SELECT state FROM info_power")
    power_list = ','.join([f'"{option["state"]}"' for option in power_options])
    power_validation = DataValidation(type="list", formula1=f'"{power_list}"', allow_blank=True)
    power_validation.error = "올바른 전원이중화 옵션을 선택하세요."
    power_validation.errorTitle = "입력 오류"
    ws.add_data_validation(power_validation)
    power_validation.add(f"P2:P1000")

    # OS 드롭다운
    os_options = execute_query("SELECT state FROM info_os")
    os_list = ','.join([f'"{option["state"]}"' for option in os_options])
    os_validation = DataValidation(type="list", formula1=f'"{os_list}"', allow_blank=True)
    os_validation.error = "올바른 OS를 선택하세요."
    os_validation.errorTitle = "입력 오류"
    ws.add_data_validation(os_validation)
    os_validation.add(f"R2:R1000")

    # 안내 시트 추가
    guide_ws = wb.create_sheet(title="작성 가이드")
    guide_ws['A1'] = "자산 일괄 등록 양식 작성 가이드"
    guide_ws['A1'].font = Font(size=14, bold=True)
    guide_ws.merge_cells('A1:F1')

    guide_rows = [
        ["* 표시된 항목은 필수 입력 항목입니다."],
        ["날짜는 YYYY-MM-DD 형식으로 입력해주세요. (예: 2023-01-01)"],
        ["드롭다운 목록에서 선택 가능한 항목만 입력해주세요."],
        ["상위자산은 그룹이 '논리'인 경우에만 입력하며, 물리 서버의 자산번호를 입력해주세요."],
        [""],
        ["각 필드 설명:"],
        ["서버명", "서버의 이름을 입력합니다."],
        ["IP 주소", "서버의 IP 주소를 입력합니다."],
        ["호스트 이름", "서버의 호스트 이름을 입력합니다."],
        ["ITAM자산번호", "ITAM 시스템의 자산 번호를 입력합니다."],
        ["센터", "서버가 위치한 센터를 입력합니다."],
        ["상면번호", "서버의 상면 번호를 입력합니다. (예: 1F-R01-01)"],
        ["상단번호", "서버의 상단 번호를 입력합니다."],
        ["도메인", "드롭다운에서 도메인을 선택합니다. (서버, 스위치, 스토리지, 어플라이언스)"],
        ["그룹", "드롭다운에서 그룹을 선택합니다. (물리, 논리, L2, L3 등)"],
        ["설치일자", "서버 설치 일자를 YYYY-MM-DD 형식으로 입력합니다."],
        ["폐기일자", "서버 폐기 일자를 YYYY-MM-DD 형식으로 입력합니다."],
        ["담당자(정)", "주 담당자 이름을 입력합니다."],
        ["담당자(부)", "부 담당자 이름을 입력합니다."],
        ["사용여부", "드롭다운에서 사용 여부를 선택합니다."],
        ["서비스구분", "드롭다운에서 서비스 구분을 선택합니다."],
        ["전원이중화", "드롭다운에서 전원 이중화 여부를 선택합니다."],
        ["PDU", "PDU 정보를 입력합니다."],
        ["OS", "드롭다운에서 운영체제를 선택합니다."],
        ["OS버전", "운영체제 버전을 입력합니다."],
        ["제조사", "서버 제조사를 입력합니다."],
        ["모델", "서버 모델명을 입력합니다."],
        ["시리얼넘버", "서버 시리얼 번호를 입력합니다."],
        ["현업담당자", "현업 담당자 이름을 입력합니다."],
        ["장비크기(U)", "서버 장비 크기를 U 단위로 입력합니다."],
        ["물리코어", "서버의 물리 코어 수를 입력합니다."],
        ["메모리(GB)", "서버의 메모리 크기를 GB 단위로 입력합니다."],
        ["상위자산", "논리 자산인 경우 상위 물리 서버의 자산 번호를 입력합니다."]
    ]

    for row_idx, row_data in enumerate(guide_rows, 2):
        for col_idx, cell_value in enumerate(row_data, 1):
            guide_ws.cell(row=row_idx, column=col_idx, value=cell_value)

    # 가이드 시트 열 너비 설정
    guide_ws.column_dimensions['A'].width = 15
    guide_ws.column_dimensions['B'].width = 50

    # 템플릿 파일 저장
    os.makedirs(os.path.dirname(template_path), exist_ok=True)
    wb.save(template_path)

    return send_file(template_path, as_attachment=True, download_name='자산_일괄등록_양식.xlsx')


@asset_bp.route('/bulk_upload', methods=['POST'])
def bulk_upload():
    """자산 일괄 등록 처리"""
    if 'excelFile' not in request.files:
        return jsonify({
            'success': False,
            'message': '업로드된 파일이 없습니다.'
        })

    file = request.files['excelFile']

    if file.filename == '':
        return jsonify({
            'success': False,
            'message': '선택된 파일이 없습니다.'
        })

    if not file.filename.endswith('.xlsx'):
        return jsonify({
            'success': False,
            'message': 'XLSX 형식의 파일만 업로드 가능합니다.'
        })

    # 임시 파일로 저장
    filename = secure_filename(file.filename)
    temp_path = os.path.join(os.getcwd(), 'uploads', filename)
    os.makedirs(os.path.dirname(temp_path), exist_ok=True)
    file.save(temp_path)

    try:
        # 엑셀 파일 로드
        wb = openpyxl.load_workbook(temp_path)
        ws = wb.active

        # 헤더 확인
        headers = [cell.value for cell in ws[1]]
        expected_headers = [
            '서버명*', 'IP 주소*', '호스트 이름', 'ITAM자산번호', '센터', '상면번호', '상단번호',
            '도메인*', '그룹*', '설치일자', '폐기일자', '담당자(정)', '담당자(부)', '사용여부*',
            '서비스구분*', '전원이중화', 'PDU', 'OS*', 'OS버전', '제조사', '모델',
            '시리얼넘버', '현업담당자', '장비크기(U)', '물리코어', '메모리(GB)', '상위자산'
        ]

        if headers != expected_headers:
            return jsonify({
                'success': False,
                'message': '양식이 올바르지 않습니다. 제공된 템플릿을 사용해주세요.'
            })

        # 데이터 추출 및 검증
        data = []
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if all(cell is None or cell == '' for cell in row):
                continue  # 빈 행 무시

            # 필수 필드 검증
            if not row[0]:  # 서버명
                errors.append({'row': row_idx, 'column': '서버명*', 'message': '서버명은 필수 입력 항목입니다.'})

            if not row[1]:  # IP 주소
                errors.append({'row': row_idx, 'column': 'IP 주소*', 'message': 'IP 주소는 필수 입력 항목입니다.'})

            # 도메인 검증
            domain_value = row[7]
            if not domain_value:
                errors.append({'row': row_idx, 'column': '도메인*', 'message': '도메인은 필수 입력 항목입니다.'})
            else:
                domain_options = [option['state'] for option in execute_query("SELECT state FROM info_domain")]
                if domain_value not in domain_options:
                    errors.append(
                        {'row': row_idx, 'column': '도메인*', 'message': f'도메인은 {", ".join(domain_options)} 중 하나여야 합니다.'})

            # 그룹 검증
            group_value = row[8]
            if not group_value:
                errors.append({'row': row_idx, 'column': '그룹*', 'message': '그룹은 필수 입력 항목입니다.'})
            else:
                group_options = [option['state'] for option in execute_query("SELECT DISTINCT state FROM info_group")]
                if group_value not in group_options:
                    errors.append({'row': row_idx, 'column': '그룹*',
                                   'message': f'그룹은 {", ".join(group_options)} 중 하나여야 합니다.'})

            # 사용여부 검증
            isoper_value = row[13]
            if not isoper_value:
                errors.append({'row': row_idx, 'column': '사용여부*', 'message': '사용여부는 필수 입력 항목입니다.'})
            else:
                isoper_options = [option['state'] for option in execute_query("SELECT state FROM info_isoper")]
                if isoper_value not in isoper_options:
                    errors.append({'row': row_idx, 'column': '사용여부*',
                                   'message': f'사용여부는 {", ".join(isoper_options)} 중 하나여야 합니다.'})

            # 서비스구분 검증
            oper_value = row[14]
            if not oper_value:
                errors.append({'row': row_idx, 'column': '서비스구분*', 'message': '서비스구분은 필수 입력 항목입니다.'})
            else:
                oper_options = [option['state'] for option in execute_query("SELECT state FROM info_oper")]
                if oper_value not in oper_options:
                    errors.append({'row': row_idx, 'column': '서비스구분*',
                                   'message': f'서비스구분은 {", ".join(oper_options)} 중 하나여야 합니다.'})

            # OS 검증
            os_value = row[17]
            if not os_value:
                errors.append({'row': row_idx, 'column': 'OS*', 'message': 'OS는 필수 입력 항목입니다.'})
            else:
                os_options = [option['state'] for option in execute_query("SELECT state FROM info_os")]
                if os_value not in os_options:
                    errors.append(
                        {'row': row_idx, 'column': 'OS*', 'message': f'OS는 {", ".join(os_options)} 중 하나여야 합니다.'})

            # 전원이중화 검증 (필수는 아님)
            power_value = row[15]
            if power_value:
                power_options = [option['state'] for option in execute_query("SELECT state FROM info_power")]
                if power_value not in power_options:
                    errors.append({'row': row_idx, 'column': '전원이중화',
                                   'message': f'전원이중화는 {", ".join(power_options)} 중 하나여야 합니다.'})

            # 날짜 형식 검증
            datein = row[9]
            if datein:
                try:
                    if isinstance(datein, str):
                        datetime.strptime(datein, '%Y-%m-%d')
                    # 엑셀 날짜 형식인 경우는 이미 datetime 객체로 로드됨
                except ValueError:
                    errors.append({'row': row_idx, 'column': '설치일자', 'message': '설치일자는 YYYY-MM-DD 형식이어야 합니다.'})

            dateout = row[10]
            if dateout:
                try:
                    if isinstance(dateout, str):
                        datetime.strptime(dateout, '%Y-%m-%d')
                except ValueError:
                    errors.append({'row': row_idx, 'column': '폐기일자', 'message': '폐기일자는 YYYY-MM-DD 형식이어야 합니다.'})

            # 상위자산 검증 (그룹이 '논리'인 경우에만)
            if group_value == '논리':
                vcenter = row[26]  # 상위자산
                if vcenter:
                    # 상위자산이 존재하는지 확인
                    sql = "SELECT COUNT(*) as count FROM total_asset WHERE pnum = %s AND `group` = 0"
                    result = execute_query(sql, (vcenter,), fetch_all=False)
                    if result['count'] == 0:
                        errors.append(
                            {'row': row_idx, 'column': '상위자산', 'message': f'상위자산 번호 {vcenter}에 해당하는 물리 서버가 존재하지 않습니다.'})

            # 데이터 추가
            data.append({
                'servername': row[0],
                'ip': row[1],
                'hostname': row[2],
                'itamnum': row[3],
                'center': row[4],
                'loc1': row[5],
                'loc2': row[6],
                'domain': row[7],
                'group': row[8],
                'datein': row[9],
                'dateout': row[10],
                'charge': row[11],
                'charge2': row[12],
                'isoper': row[13],
                'oper': row[14],
                'power': row[15],
                'pdu': row[16],
                'os': row[17],
                'osver': row[18],
                'maker': row[19],
                'model': row[20],
                'serial': row[21],
                'charge3': row[22],
                'usize': row[23],
                'cpucore': row[24],
                'memory': row[25],
                'vcenter': row[26] if row[8] == '논리' else None
            })

        # 오류가 있으면 처리 중단
        if errors:
            return jsonify({
                'success': False,
                'message': f'{len(errors)}개의 오류가 발견되었습니다. 수정 후 다시 시도해주세요.',
                'errors': errors
            })

        # 데이터 DB에 저장
        inserted_count = 0
        db = get_db_connection()
        try:
            with db.cursor() as cursor:
                for item in data:
                    # 코드값 조회
                    # domain
                    cursor.execute("SELECT domain FROM info_domain WHERE state = %s", (item['domain'],))
                    domain_value = cursor.fetchone()
                    domain = domain_value['domain'] if domain_value else 0

                    # group
                    cursor.execute("SELECT `group` FROM info_group WHERE state = %s AND domain = %s",
                                   (item['group'], domain))
                    group_value = cursor.fetchone()
                    group = group_value['group'] if group_value else 0

                    # isoper
                    cursor.execute("SELECT isoper FROM info_isoper WHERE state = %s", (item['isoper'],))
                    isoper_value = cursor.fetchone()
                    isoper = isoper_value['isoper'] if isoper_value else 0

                    # oper
                    cursor.execute("SELECT oper FROM info_oper WHERE state = %s", (item['oper'],))
                    oper_value = cursor.fetchone()
                    oper = oper_value['oper'] if oper_value else 0

                    # power
                    power = None
                    if item['power']:
                        cursor.execute("SELECT power FROM info_power WHERE state = %s", (item['power'],))
                        power_value = cursor.fetchone()
                        power = power_value['power'] if power_value else None

                    # os
                    cursor.execute("SELECT os FROM info_os WHERE state = %s", (item['os'],))
                    os_value = cursor.fetchone()
                    os_code = os_value['os'] if os_value else 0

                    # 날짜 처리
                    datein = item['datein']
                    if isinstance(datein, str) and datein:
                        datein = datetime.strptime(datein, '%Y-%m-%d').date()
                    elif isinstance(datein, datetime):
                        datein = datein.date()

                    dateout = item['dateout']
                    if isinstance(dateout, str) and dateout:
                        dateout = datetime.strptime(dateout, '%Y-%m-%d').date()
                    elif isinstance(dateout, datetime):
                        dateout = dateout.date()

                    # 숫자 필드 처리
                    loc2 = int(item['loc2']) if item['loc2'] else None
                    usize = int(item['usize']) if item['usize'] else 1
                    cpucore = int(item['cpucore']) if item['cpucore'] else None
                    memory = int(item['memory']) if item['memory'] else None
                    vcenter = int(item['vcenter']) if item['vcenter'] else None

                    # 현재 시간
                    now = datetime.now()

                    # SQL 쿼리
                    sql = """INSERT INTO total_asset (
                        itamnum, servername, ip, hostname, center, loc1, loc2, `group`, vcenter, 
                        datein, dateout, charge, charge2, isoper, oper, power, pdu, os, osver, 
                        maker, model, serial, domain, charge3, usize, cpucore, memory, 
                        isfix, dateinsert, dateupdate
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                    )"""

                    cursor.execute(sql, (
                        item['itamnum'], item['servername'], item['ip'], item['hostname'],
                        item['center'], item['loc1'], loc2, group, vcenter,
                        datein, dateout, item['charge'], item['charge2'],
                        isoper, oper, power, item['pdu'], os_code, item['osver'],
                        item['maker'], item['model'], item['serial'], domain, item['charge3'],
                        usize, cpucore, memory, 1, now, now
                    ))

                    inserted_count += 1

                db.commit()
        except Exception as e:
            db.rollback()
            return jsonify({
                'success': False,
                'message': f'데이터 저장 중 오류가 발생했습니다: {str(e)}'
            })
        finally:
            db.close()

        # ���시 무효화
        invalidate_cache_pattern('index_data')
        invalidate_cache_pattern('asset_data')
        invalidate_cache_pattern('asset_graph')

        return jsonify({
            'success': True,
            'message': f'{inserted_count}개의 자산이 성공적으로 등록되었습니다.',
            'details': {
                '총 등록 자산': inserted_count,
                '물리 서버': sum(1 for item in data if item['group'] == '물리'),
                '논리 서버': sum(1 for item in data if item['group'] == '논리')
            }
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'파일 처리 중 오류가 발생했습니다: {str(e)}'
        })
    finally:
        # 임시 파일 삭제
        if os.path.exists(temp_path):
            os.remove(temp_path)
