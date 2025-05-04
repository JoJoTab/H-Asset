from flask import Blueprint, render_template, redirect, url_for, request, send_file, jsonify
import pandas as pd
import openpyxl
import os
from utils.db import execute_query
from utils.cache import cache, invalidate_cache_pattern

rack_bp = Blueprint('rack', __name__)


@rack_bp.route('/racklayout', methods=['GET'])
@cache('racklayout')
def racklayout():
    """랙 레이아웃 페이지"""
    sql = "SELECT loc1 FROM total_asset WHERE `group` = 0"
    data = execute_query(sql)

    # 데이터프레임 변환
    columns = [column for column in data[0].keys()] if data else []
    df = pd.DataFrame(data, columns=columns)
    df = df.dropna(subset=['loc1'])

    df['loc1'] = df['loc1'].apply(lambda x: x[:-4] + x[-3:-2] + '-' + x[-2:])

    floors = set()
    floor_columns = {}

    for loc in df['loc1']:
        parts = loc.split('-')
        if len(parts) < 3:
            continue  # 형식이 올바르지 않으면 건너뜀

        floor = parts[0]
        column = parts[1]  # R01, L01 등

        floors.add(floor)
        if floor not in floor_columns:
            floor_columns[floor] = set()
        floor_columns[floor].add(column)

    # 랙 정보 가져오기
    sql_rack_info = "SELECT loc, rackname, rackenable FROM rack_info"
    rack_info_data = execute_query(sql_rack_info)

    # 데이터프레임 변환
    columns = [column for column in rack_info_data[0].keys()] if rack_info_data else []
    df_rack = pd.DataFrame(rack_info_data, columns=columns)
    df_rack['loc'] = df_rack['loc'].apply(lambda x: x[:-4] + x[-3:-2] + '-' + x[-2:])
    rack_info_dict = {row['loc']: (row['rackname'], row['rackenable']) for _, row in df_rack.iterrows()}

    # 클릭 가능한 슬롯 생성
    equipment_data = {}
    for floor in floors:
        for column in floor_columns[floor]:
            for loc in range(0, 16):  # 00~15 (0부터 시작)
                loc_string = '{0:02d}'.format(loc)
                loc_key = "{}-{}-{}".format(floor, column, loc_string)
                equipment_data[loc_key] = loc_key in df['loc1'].values
        if floor in floor_columns:
            # R과 L로 분리
            r_items = [item for item in floor_columns[floor] if item.endswith('R')]
            l_items = [item for item in floor_columns[floor] if item.endswith('L')]

            # R과 L 각각 정렬
            r_items.sort(key=lambda x: (int(x[:-1]), x))  # 숫자 오름차순 정렬
            l_items.sort(key=lambda x: (int(x[:-1]), x))  # 숫자 오름차순 정렬

            sorted_items = r_items + l_items
            floor_columns[floor] = sorted_items

    return render_template('racklayout.html', equipment_data=equipment_data, floors=sorted(floors),
                           floor_columns=floor_columns, rack_info_dict=rack_info_dict)


@rack_bp.route('/racklayout_edit', methods=['GET', 'POST'])
def racklayout_edit():
    """랙 레이아웃 편집 페이지"""
    if request.method == 'POST':
        loc = request.form.get('update')  # 수정 버튼을 통해 전달된 loc 값
        if loc:
            rackname = request.form.get(f'rackname_{loc}', '')
            rackenable = int(request.form.get(f'rackenable_{loc}', 1))

            # loc에 해당하는 rack_info 업데이트 또는 삽입
            sql_check = "SELECT COUNT(*) as count FROM rack_info WHERE loc = %s"
            result = execute_query(sql_check, (loc,), fetch_all=False)
            exists = result['count'] if result else 0

            if exists > 0:
                # 존재하면 업데이트
                sql_update = """
               UPDATE rack_info 
               SET rackname = %s, rackenable = %s 
               WHERE loc = %s
               """
                execute_query(sql_update, (rackname, rackenable, loc), fetch_all=False)
            else:
                # 존재하지 않으면 삽입
                sql_insert = """
               INSERT INTO rack_info (loc, rackname, rackenable) 
               VALUES (%s, %s, %s)
               """
                execute_query(sql_insert, (loc, rackname, rackenable), fetch_all=False)

            # 캐시 무효화
            invalidate_cache_pattern('racklayout')

    # loc1 정보 조회
    sql_loc = "SELECT DISTINCT loc1 FROM total_asset"
    loc_data = execute_query(sql_loc)

    # loc_data에서 loc1 값 추출 및 정렬
    loc_list = sorted([row['loc1'] for row in loc_data if row['loc1'] is not None])

    sql_select = "SELECT loc, rackname, rackenable FROM rack_info WHERE loc IN %s"
    current_data = execute_query(sql_select, (tuple(loc_list),))

    # 데이터 사전 생성
    current_dict = {row['loc']: (row['rackname'], row['rackenable']) for row in current_data}

    return render_template('racklayout_edit.html', loc_list=loc_list, current_dict=current_dict)


@rack_bp.route('/rackview', methods=['GET', 'POST'])
def rackview():
    """랙 뷰 페이지"""
    sql = "SELECT pnum, loc1, loc2, servername, charge, maker, model, usize FROM total_asset"
    data = execute_query(sql)

    # 데이터프레임 변환
    columns = [column for column in data[0].keys()] if data else []
    df = pd.DataFrame(data, columns=columns)

    # loc1의 값이 없는 행 제거
    df = df.dropna(subset=['loc1'])

    # loc 컬럼에서 층, 열, 위치 추출
    floors = set()
    columns_set = set()
    locations = set()

    for loc in df['loc1']:
        try:
            floor, column, location = loc.split('-')
            floors.add(floor)
            columns_set.add(column)
            locations.add(location)
        except ValueError:
            continue

    arg_floor = request.args.get('floor')
    arg_column = request.args.get('column')
    arg_location = request.args.get('location')

    if request.method == 'POST':
        selected_floor = request.form.get('floor')
        selected_column = request.form.get('column')
        selected_location = request.form.get('location')

        # loc 필터링
        filtered_df = df[df['loc1'] == f"{selected_floor}-{selected_column}-{selected_location}"]

        # 결과 저장할 리스트
        result_df = pd.DataFrame()

        # 상단번호 42부터 1까지 반복
        idx = 42
        while idx > 0:
            loc2_rows = filtered_df[filtered_df['loc2'] == idx]
            if not loc2_rows.empty:
                # 중복된 모델명, 서버명, 담당자, pnum 수집
                model_names = loc2_rows['model'].dropna().unique().tolist()
                server_names = loc2_rows['servername'].dropna().unique().tolist()
                charges = loc2_rows['charge'].dropna().unique().tolist()
                maker_names = loc2_rows['maker'].dropna().unique().tolist()
                pnums = loc2_rows['pnum'].dropna().unique().tolist()

                # 각 정보 합치기 (줄바꿈으로 구분)
                maker_combined = '<br>'.join(filter(None, maker_names))
                model_combined = '<br>'.join(filter(None, model_names))
                server_combined = '<br>'.join(filter(None, server_names))
                charge_combined = '<br>'.join(filter(None, charges))
                pnum_combined = ','.join(map(str, filter(None, pnums)))

                # usize로 병합할 행 수 결정
                usize = int(loc2_rows['usize'].max())
                result_df[idx] = dict(maker=maker_combined, model=model_combined,
                                      servername=server_combined, charge=charge_combined,
                                      usize=usize, pnum=pnum_combined)
                idx -= usize - 1
            else:
                result_df[idx] = dict(maker='', model='', servername='', charge='', usize=1, pnum='')
            idx -= 1

        temp_result = []
        for idx in result_df.columns:
            temp_result.append({'loc2': idx, 'maker': result_df[idx].maker, 'model': result_df[idx].model,
                                'servername': result_df[idx].servername, 'charge': result_df[idx].charge,
                                'usize': result_df[idx].usize, 'pnum': result_df[idx].pnum})

        return jsonify(temp_result)

    # 상면번호 내림차순 정렬
    equipment_data = [{'id': i, 'name': '', 'owner': ''} for i in range(42, 0, -1)]

    # 모든 값이 존재하는지 확인
    if arg_floor and arg_column and arg_location:
        return render_template('rack.html',
                               equipment_data=equipment_data,
                               floors=list(floors),
                               columns=list(columns_set),
                               locations=list(locations),
                               selected_floor=arg_floor,
                               selected_column=arg_column,
                               selected_location=arg_location,
                               auto_fetch=True)

    return render_template('rack.html',
                           equipment_data=equipment_data,
                           floors=list(floors),
                           columns=list(columns_set),
                           locations=list(locations))


@rack_bp.route('/get_locations', methods=['GET'])
def get_locations():
    """위치 정보 가져오기"""
    floor = request.args.get('floor')
    column = request.args.get('column')

    sql = "SELECT loc1 FROM total_asset WHERE loc1 LIKE %s"
    data = execute_query(sql, (f"{floor}-{column}-%",))

    # loc1에서 위치만 추출
    locations = set(loc['loc1'].split('-')[2] for loc in data)

    return jsonify(sorted(list(locations)))


@rack_bp.route('/get_columns', methods=['GET'])
def get_columns():
    """컬럼 정보 가져오기"""
    floor = request.args.get('floor')

    sql = "SELECT DISTINCT loc1 FROM total_asset WHERE loc1 LIKE %s"
    data = execute_query(sql, (f"{floor}-%",))

    # loc1에서 column만 추출
    columns = set(loc['loc1'].split('-')[1] for loc in data)

    return jsonify(sorted(list(columns)))


@rack_bp.route('/rack_export', methods=['GET', 'POST'])
def rack_export():
    """랙 정보 내보내기"""
    # 현재 테이블 데이터 가져오기
    sql = "SELECT loc1, loc2, servername, charge, maker, model, usize FROM total_asset"
    data = execute_query(sql)

    # 데이터프레임 변환
    columns = [column for column in data[0].keys()] if data else []
    df = pd.DataFrame(data, columns=columns)

    # loc1의 값이 없는 행 제거
    df = df.dropna(subset=['loc1'])

    selected_floor = request.form.get('floor')
    selected_column = request.form.get('column')
    selected_location = request.form.get('location')

    filtered_df = df[df['loc1'] == f"{selected_floor}-{selected_column}-{selected_location}"]
    filtered_df = filtered_df[['loc2', 'model', 'servername', 'charge', 'usize']]

    # loc2 범위 설정
    loc2_range = range(1, 43)

    # 현재 loc2 값
    existing_loc2 = filtered_df['loc2'].unique()

    # 추가할 데이터프레임 생성
    new_rows = []

    # loc2가 없는 값에 대한 데이터 추가
    for loc2 in loc2_range:
        if loc2 not in existing_loc2:
            # loc2에 해당하는 데이터 추가
            new_row = {'loc2': loc2, 'model': '', 'servername': '', 'charge': '', 'usize': ''}
            new_rows.append(new_row)

    # 새로운 행들을 데이터프레임으로 변환
    new_rows_df = pd.DataFrame(new_rows)

    # 기존 데이터프레임과 새로운 행들을 결합
    filtered_df = pd.concat([filtered_df, new_rows_df], ignore_index=True)

    # loc2 기준으로 내림차순 정렬
    filtered_df.sort_values(by='loc2', ascending=False, inplace=True)
    filtered_df.reset_index(drop=True, inplace=True)

    template_path = 'static/excel/template_rack.xlsx'
    # 엑셀 템플릿 파일 열기
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active  # 기본 시트 선택

    # 데이터 삽입 시작 위치 설정
    start_row = 4
    start_col = 2

    sheet.cell(row=2, column=2).value = f"{selected_floor}-{selected_column}-{selected_location}"

    # 데이터 삽입
    for loc2 in filtered_df['loc2'].unique():  # loc2의 고유 값으로 반복
        rows = filtered_df[filtered_df['loc2'] == loc2]  # loc2에 해당하는 모든 행 선택

        for i, row in rows.iterrows():  # 각 행을 반복
            for j, col in enumerate(filtered_df.columns):  # 열 인덱스 사용
                if j == 0: continue
                cell = sheet.cell(row=start_row + loc2, column=start_col + j)

                # 기존 값이 있을 경우 개행하여 추가
                if cell.value:
                    # cell.value가 int인 경우 문자열로 변환
                    if isinstance(cell.value, int):
                        cell.value = str(cell.value)
                    cell.value += f"\n{row[col]}"  # 여기서 row[col]을 추가
                else:
                    cell.value = row[col]  # 첫 번째 값 설정

    # 셀의 높이를 자동 조정하기 위해 개행 설정
    for row in sheet.iter_rows(min_row=start_row, max_row=start_row + len(filtered_df) - 1, min_col=start_col,
                               max_col=start_col + len(filtered_df.columns) - 1):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    # 변경된 엑셀 파일 저장
    output_path = 'output.xlsx'
    workbook.save(output_path)

    return send_file(output_path, as_attachment=True)
